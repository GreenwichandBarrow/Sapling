from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASELINE = Path("/home/ubuntu/projects/Sapling/brain/operating-areas/c-suite/CIO/G&B Financial Model Template.xlsx")
OUTPUT = Path("/home/ubuntu/projects/Sapling/brain/operating-areas/c-suite/CIO/Pest Multi-Acquisition Financial Model - First Pass 2026-06-18.xlsx")


BLUE = "1F4E79"
LIGHT_BLUE = "D9EAF7"
YELLOW = "FFF2CC"
GREEN = "E2F0D9"
RED = "FCE4D6"
GREY = "E7E6E6"
WHITE = "FFFFFF"
BLACK = "000000"


def header(cell, text, fill=BLUE):
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, color=WHITE if fill == BLUE else BLACK)
    cell.alignment = Alignment(horizontal="center")


def label(cell, text):
    cell.value = text
    cell.font = Font(bold=True)


def note(cell, text):
    cell.value = text
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def input_cell(cell, value):
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=YELLOW)


def formula(cell, value):
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=GREEN)


def money_fmt(ws, refs):
    for ref in refs:
        for row in ws[ref]:
            for cell in row:
                cell.number_format = '$#,##0;[Red]($#,##0);-'


def pct_fmt(ws, refs):
    for ref in refs:
        for row in ws[ref]:
            for cell in row:
                cell.number_format = '0.0%'


def mult_fmt(ws, refs):
    for ref in refs:
        for row in ws[ref]:
            for cell in row:
                cell.number_format = '0.0x'


def style_sheet(ws):
    thin = Side(style="thin", color="D9E1F2")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = copy(cell.alignment) if cell.alignment else Alignment()
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical="center",
                wrap_text=cell.alignment.wrap_text,
            )
    ws.freeze_panes = "B7"


def setup_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width



def build_pest_single_acquisition(wb):
    src = wb["Model Template"]
    ws = wb.copy_worksheet(src)
    ws.title = "Pest Single Acquisition"
    ws["B2"] = "G&B Model - Pest Single Acquisition Base Case"
    ws["B12"] = "$000s unless otherwise indicated"

    # Transaction assumptions: keep the G&B template layout; replace only yellow/input assumptions.
    ws["N15"] = 2500  # Prior offer / TEV proxy from Melissa call, not valuation conclusion.
    ws["N16"] = "EBITDA"
    ws["N18"] = "2026-12-31"
    ws["N19"] = 0.05
    ws["N20"] = 0  # Melissa: no bank debt expected at this size.
    ws["N21"] = 750  # 30% seller note on $2.5M proxy.
    ws["N22"] = 200
    ws["N23"] = 0
    ws["N24"] = "C-corp"
    ws["N25"] = 0.125
    ws["N26"] = 0
    ws["N27"] = 0
    ws["N28"] = 0

    # Historical/current financial inputs. Older periods are intentionally blank/TBD rather than invented.
    ws["C37"] = 0
    ws["D37"] = 0
    ws["E37"] = 1500
    ws["F38"] = 0.05
    ws["G38"] = 0.05
    ws["H38"] = 0.05
    ws["I38"] = 0.05
    ws["J38"] = 0.05
    ws["C40"] = 0
    ws["D40"] = 0
    ws["E40"] = 500
    ws["F41"] = 0.3333333333
    ws["G41"] = 0.3333333333
    ws["H41"] = 0.3333333333
    ws["I41"] = 0.3333333333
    ws["J41"] = 0.3333333333

    # Debt schedule assumptions from Melissa's range.
    ws["L80"] = 0.00
    ws["L89"] = 0.07
    # Seller note amortization across six years: template has five forecast years, so show interest pressure and flag final-year balloon externally.
    ws["F91"] = 0
    ws["G91"] = 0
    ws["H91"] = 0
    ws["I91"] = 0
    ws["J91"] = 1

    ws["L31"] = "Assumptions & Source Notes"
    ws["L33"] = "Melissa 6/17: core $1.5M revenue / $500K EBITDA; 5% organic growth; equity + seller note, no bank debt at close."
    ws["L34"] = "Seller note base: 30% of TEV proxy, 7% interest. Rollover base: 12.5%."
    ws["L35"] = "Historical C/D periods are set to 0/TBD because source material did not provide them."
    return ws

def build_base_case(wb):
    ws = wb.create_sheet("Pest Base Case", 0)
    setup_widths(ws, {
        "A": 3, "B": 34, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14,
        "H": 14, "I": 14, "J": 14, "K": 14, "L": 14, "M": 18, "N": 44
    })

    ws["B2"] = "G&B Pest Multi-Acquisition Model - First Pass"
    ws["B2"].font = Font(bold=True, size=16, color=BLUE)
    ws["B3"] = "Built from G&B Financial Model Template.xlsx. $000s unless otherwise indicated. Yellow cells are inputs; green cells are formulas."
    ws["B3"].font = Font(italic=True, color="666666")

    years = ["Close", "Y1", "Y2", "Y3", "Y4", "Y5"]
    for idx, year in enumerate(years, start=4):
        header(ws.cell(6, idx), year)
    header(ws["M6"], "Source / status")
    header(ws["N6"], "Notes")

    label(ws["B8"], "Core business assumptions")
    inputs = [
        ("Revenue", 1500, "Melissa call 2026-06-17", "Core business: $1.5M revenue."),
        ("EBITDA", 500, "Melissa call 2026-06-17", "Core business: $500K EBITDA, 33.3% margin."),
        ("Labor / COGS % revenue", 0.40, "Melissa call 2026-06-17", "40% labor / COGS."),
        ("Organic growth", 0.05, "Melissa call 2026-06-17", "Melissa recommended ~5% organic growth."),
        ("Initial tech count", 6, "Guillermo/Carlos 2026-06-17", "6-person team; technician count still needs diligence."),
        ("Revenue / tech proxy", "=D9/D13", "Formula", "Useful only after technician count is verified."),
    ]
    start = 9
    for r, (name, val, src, notes) in enumerate(inputs, start=start):
        ws.cell(r, 2).value = name
        if isinstance(val, str) and val.startswith("="):
            formula(ws.cell(r, 4), val)
        else:
            input_cell(ws.cell(r, 4), val)
        ws.cell(r, 13).value = src
        note(ws.cell(r, 14), notes)

    label(ws["B17"], "Transaction assumptions")
    txn = [
        ("Prior offer / TEV proxy", 2500, "Melissa call 2026-06-17", "Seller reportedly had $2.5M offer last year; use as proxy, not valuation conclusion."),
        ("Entry multiple proxy", "=D18/D10", "Formula", "Implied by prior-offer proxy and $500K EBITDA."),
        ("Transaction costs", 0.05, "Baseline workbook", "Carryover placeholder from single-acquisition model; diligence needed."),
        ("Cash to balance sheet", 200, "Baseline workbook", "Placeholder from baseline workbook; adjust after operating cash need is known."),
        ("Seller note % of TEV", 0.30, "Model input / Melissa structure", "Structure lever; stress-tested on Stress Tests tab."),
        ("Seller note interest rate", 0.07, "Melissa call 2026-06-17", "Melissa range 6-8%; base at midpoint."),
        ("Seller note tenor", 6, "Melissa call 2026-06-17", "Melissa range 5-7 years; base at midpoint."),
        ("Rollover equity %", 0.125, "Jay + Melissa 2026-06-17", "Jay suggested 10-15% max; Melissa said minority governance risk is low."),
        ("Term debt at close", 0, "Melissa call 2026-06-17", "No bank debt expected at this size."),
        ("Bank debt threshold EBITDA", 2000, "Melissa call 2026-06-17", "~$2M EBITDA minimum for bank debt."),
        ("Indicative bank facility", 5000, "Melissa call 2026-06-17", "~$5M facility once threshold is reached."),
    ]
    for r, (name, val, src, notes) in enumerate(txn, start=18):
        ws.cell(r, 2).value = name
        if isinstance(val, str) and val.startswith("="):
            formula(ws.cell(r, 4), val)
        else:
            input_cell(ws.cell(r, 4), val)
        ws.cell(r, 13).value = src
        note(ws.cell(r, 14), notes)

    label(ws["B31"], "Sources & uses")
    su = [
        ("TEV", "=D18"),
        ("Transaction costs", "=D18*D20"),
        ("Cash to balance sheet", "=D21"),
        ("Total uses", "=SUM(D32:D34)"),
        ("Seller note", "=D18*D22"),
        ("Seller rollover", "=D18*D25"),
        ("Term debt", "=D26"),
        ("Investor equity required", "=D35-D36-D37-D38"),
        ("Check", "=D35-SUM(D36:D39)"),
    ]
    for r, (name, val) in enumerate(su, start=32):
        ws.cell(r, 2).value = name
        formula(ws.cell(r, 4), val)
    ws["M32"] = "Formula"; ws["N32"] = "Capital stack is intentionally simple for the first pass."

    label(ws["B43"], "Operating forecast")
    for idx, year in enumerate(years, start=4):
        header(ws.cell(44, idx), year, fill=GREY)
    rows = [
        ("Core revenue", "=D9", "=D45*(1+$D$12)", "=E45*(1+$D$12)", "=F45*(1+$D$12)", "=G45*(1+$D$12)", "=H45*(1+$D$12)"),
        ("Acquired revenue", "='M&A Schedule'!D22", "='M&A Schedule'!E22", "='M&A Schedule'!F22", "='M&A Schedule'!G22", "='M&A Schedule'!H22", "='M&A Schedule'!I22"),
        ("Total revenue", "=SUM(D45:D46)", "=SUM(E45:E46)", "=SUM(F45:F46)", "=SUM(G45:G46)", "=SUM(H45:H46)", "=SUM(I45:I46)"),
        ("Core EBITDA", "=D10", "=D48*(1+$D$12)", "=E48*(1+$D$12)", "=F48*(1+$D$12)", "=G48*(1+$D$12)", "=H48*(1+$D$12)"),
        ("Acquired EBITDA", "='M&A Schedule'!D46", "='M&A Schedule'!E46", "='M&A Schedule'!F46", "='M&A Schedule'!G46", "='M&A Schedule'!H46", "='M&A Schedule'!I46"),
        ("Corporate costs", "=SUM(D57:D60)", "=SUM(E57:E60)", "=SUM(F57:F60)", "=SUM(G57:G60)", "=SUM(H57:H60)", "=SUM(I57:I60)"),
        ("Adjusted EBITDA", "=SUM(D48:D49)-D50", "=SUM(E48:E49)-E50", "=SUM(F48:F49)-F50", "=SUM(G48:G49)-G50", "=SUM(H48:H49)-H50", "=SUM(I48:I49)-I50"),
        ("EBITDA margin", "=D51/D47", "=E51/E47", "=F51/F47", "=G51/G47", "=H51/H47", "=I51/I47"),
    ]
    for r, row in enumerate(rows, start=45):
        ws.cell(r, 2).value = row[0]
        for c, val in enumerate(row[1:], start=4):
            formula(ws.cell(r, c), val)

    label(ws["B56"], "Corporate cost build")
    corp = [
        ("Kay salary / CEO comp", 150, 150, 175, 200, 225, 250, "Input needed; placeholder for cash-flow pressure only."),
        ("GM / ops lead", 0, 0, 125, 150, 175, 200, "Layered future leadership per Melissa guidance."),
        ("Accounting / software / admin", 50, 60, 75, 90, 110, 130, "Shared back-end cost pool; verify with operator quotes."),
        ("Other corporate costs", 0, 25, 50, 75, 100, 125, "Placeholder for holding-company overhead."),
    ]
    for r, row in enumerate(corp, start=57):
        ws.cell(r, 2).value = row[0]
        for c, val in enumerate(row[1:7], start=4):
            input_cell(ws.cell(r, c), val)
        note(ws.cell(r, 14), row[7])

    label(ws["B63"], "Cash flow / debt-service pressure")
    cf_rows = [
        ("Adjusted EBITDA", "=D51", "=E51", "=F51", "=G51", "=H51", "=I51"),
        ("Estimated capex", "=D47*1.0%", "=E47*1.0%", "=F47*1.0%", "=G47*1.0%", "=H47*1.0%", "=I47*1.0%"),
        ("Cash taxes placeholder", "=MAX(D51-D68,0)*26.5%", "=MAX(E51-E68,0)*26.5%", "=MAX(F51-F68,0)*26.5%", "=MAX(G51-G68,0)*26.5%", "=MAX(H51-H68,0)*26.5%", "=MAX(I51-I68,0)*26.5%"),
        ("Seller note interest", "=(D36)*$D$23", "=(D36)*$D$23", "=(D36)*$D$23", "=(D36)*$D$23", "=(D36)*$D$23", "=(D36)*$D$23"),
        ("Seller note amortization", "=D36/$D$24", "=D36/$D$24", "=D36/$D$24", "=D36/$D$24", "=D36/$D$24", "=D36/$D$24"),
        ("Term debt service", "=0", "=0", "=0", "=0", "=0", "=0"),
        ("Cash after debt service", "=D64-D65-D66-D67-D68-D69", "=E64-E65-E66-E67-E68-E69", "=F64-F65-F66-F67-F68-F69", "=G64-G65-G66-G67-G68-G69", "=H64-H65-H66-H67-H68-H69", "=I64-I65-I66-I67-I68-I69"),
        ("Debt service coverage", "=D64/SUM(D67:D69)", "=E64/SUM(E67:E69)", "=F64/SUM(F67:F69)", "=G64/SUM(G67:G69)", "=H64/SUM(H67:H69)", "=I64/SUM(I67:I69)"),
        ("Bank debt eligible?", '=IF(D51>=$D$27,"Yes","No")', '=IF(E51>=$D$27,"Yes","No")', '=IF(F51>=$D$27,"Yes","No")', '=IF(G51>=$D$27,"Yes","No")', '=IF(H51>=$D$27,"Yes","No")', '=IF(I51>=$D$27,"Yes","No")'),
    ]
    for r, row in enumerate(cf_rows, start=64):
        ws.cell(r, 2).value = row[0]
        for c, val in enumerate(row[1:], start=4):
            formula(ws.cell(r, c), val)

    label(ws["B76"], "CIO verdict logic")
    ws["B77"] = "Pest path supportability"
    ws["D77"] = '=IF(MIN(E70:I70)<0,"Does not support base structure",IF(MIN(E71:I71)<1.25,"Marginal debt-service cushion","Supports first-pass structure"))'
    ws["D77"].fill = PatternFill("solid", fgColor=GREEN)
    ws["B78"] = "Buy-box fit"
    ws["D78"] = "Good niche / undersized platform: economics fit, but initial EBITDA is below preferred $750K+ floor. Needs multi-acquisition path or organic growth proof."
    ws["D78"].alignment = Alignment(wrap_text=True)

    money_fmt(ws, ["D9:I10", "D18:D18", "D21:D21", "D27:D28", "D32:D40", "D45:I51", "D57:I70"])
    pct_fmt(ws, ["D11:D12", "D20:D20", "D22:D23", "D25:D25", "D52:I52"])
    mult_fmt(ws, ["D19:D19"])
    style_sheet(ws)
    return ws


def build_ma_schedule(wb):
    ws = wb.create_sheet("M&A Schedule", 1)
    setup_widths(ws, {"A": 3, "B": 30, "C": 16, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 18, "K": 50})
    ws["B2"] = "M&A Schedule"
    ws["B2"].font = Font(bold=True, size=16, color=BLUE)
    ws["B3"] = "Acquisition cohorts use Melissa's first-pass structure. Inputs are deliberately average profiles, not target-level diligence."

    years = ["Close", "Y1", "Y2", "Y3", "Y4", "Y5"]
    for idx, year in enumerate(years, start=4):
        header(ws.cell(6, idx), year)
    header(ws["J6"], "Source / status")
    header(ws["K6"], "Notes")

    label(ws["B8"], "Acquisition count by cohort")
    cohorts = [
        ("Core platform", 1, 0, 0, 0, 0, 0, "Melissa / live target", "Existing $1.5M revenue, $500K EBITDA business."),
        ("Aquihire / small book", 0, 1, 2, 2, 2, 2, "Melissa example", "$500K revenue, $100K EBITDA, 2 technicians."),
        ("Platform tuck-in", 0, 0, 0, 1, 1, 1, "TBD", "Placeholder count only; profile needs real deal-flow evidence."),
        ("De novo crews", 0, 0, 0, 0, 0, 0, "Lower priority", "Melissa reframed as organic growth; aquihire preferred."),
    ]
    for r, row in enumerate(cohorts, start=9):
        ws.cell(r, 2).value = row[0]
        for c, val in enumerate(row[1:7], start=4):
            input_cell(ws.cell(r, c), val)
        ws.cell(r, 10).value = row[7]
        note(ws.cell(r, 11), row[8])

    label(ws["B17"], "Annual acquired revenue contribution")
    rev_rows = [
        ("Core platform", "=D9*'Pest Base Case'!$D$9", "=E9*'Pest Base Case'!$D$9", "=F9*'Pest Base Case'!$D$9", "=G9*'Pest Base Case'!$D$9", "=H9*'Pest Base Case'!$D$9", "=I9*'Pest Base Case'!$D$9"),
        ("Aquihire / small book", "=SUM($D10:D10)*$D$29", "=SUM($D10:E10)*$D$29", "=SUM($D10:F10)*$D$29", "=SUM($D10:G10)*$D$29", "=SUM($D10:H10)*$D$29", "=SUM($D10:I10)*$D$29"),
        ("Platform tuck-in", "=SUM($D11:D11)*$D$30", "=SUM($D11:E11)*$D$30", "=SUM($D11:F11)*$D$30", "=SUM($D11:G11)*$D$30", "=SUM($D11:H11)*$D$30", "=SUM($D11:I11)*$D$30"),
        ("De novo crews", "=SUM($D12:D12)*$D$31", "=SUM($D12:E12)*$D$31", "=SUM($D12:F12)*$D$31", "=SUM($D12:G12)*$D$31", "=SUM($D12:H12)*$D$31", "=SUM($D12:I12)*$D$31"),
        ("Total acquired revenue", "=SUM(D19:D21)", "=SUM(E19:E21)", "=SUM(F19:F21)", "=SUM(G19:G21)", "=SUM(H19:H21)", "=SUM(I19:I21)"),
    ]
    for r, row in enumerate(rev_rows, start=18):
        ws.cell(r, 2).value = row[0]
        for c, val in enumerate(row[1:], start=4):
            formula(ws.cell(r, c), val)

    label(ws["B26"], "Average acquisition profile")
    profile = [
        ("Core platform revenue", 1500, "Melissa call"),
        ("Core platform EBITDA", 500, "Melissa call"),
        ("Aquihire revenue", 500, "Melissa example"),
        ("Platform tuck-in revenue", 1500, "TBD / placeholder"),
        ("De novo crew revenue", 0, "Not modeled"),
        ("Aquihire EBITDA", 100, "Melissa example"),
        ("Platform tuck-in EBITDA", 300, "TBD / placeholder"),
        ("De novo crew EBITDA", 0, "Not modeled"),
        ("Aquihire techs", 2, "Melissa example"),
    ]
    for r, (name, val, src) in enumerate(profile, start=27):
        ws.cell(r, 2).value = name
        input_cell(ws.cell(r, 4), val)
        ws.cell(r, 10).value = src

    label(ws["B41"], "Annual acquired EBITDA contribution")
    ebitda_rows = [
        ("Core platform", "=D9*$D$28", "=E9*$D$28", "=F9*$D$28", "=G9*$D$28", "=H9*$D$28", "=I9*$D$28"),
        ("Aquihire / small book", "=SUM($D10:D10)*$D$32", "=SUM($D10:E10)*$D$32", "=SUM($D10:F10)*$D$32", "=SUM($D10:G10)*$D$32", "=SUM($D10:H10)*$D$32", "=SUM($D10:I10)*$D$32"),
        ("Platform tuck-in", "=SUM($D11:D11)*$D$33", "=SUM($D11:E11)*$D$33", "=SUM($D11:F11)*$D$33", "=SUM($D11:G11)*$D$33", "=SUM($D11:H11)*$D$33", "=SUM($D11:I11)*$D$33"),
        ("De novo crews", "=SUM($D12:D12)*$D$34", "=SUM($D12:E12)*$D$34", "=SUM($D12:F12)*$D$34", "=SUM($D12:G12)*$D$34", "=SUM($D12:H12)*$D$34", "=SUM($D12:I12)*$D$34"),
        ("Total acquired EBITDA", "=SUM(D43:D45)", "=SUM(E43:E45)", "=SUM(F43:F45)", "=SUM(G43:G45)", "=SUM(H43:H45)", "=SUM(I43:I45)"),
    ]
    for r, row in enumerate(ebitda_rows, start=42):
        ws.cell(r, 2).value = row[0]
        for c, val in enumerate(row[1:], start=4):
            formula(ws.cell(r, c), val)

    label(ws["B50"], "Acquisition cash need")
    cash_rows = [
        ("Aquihire TEV multiple", 3.0, "TBD", "Placeholder only. Replace with broker / owner evidence."),
        ("Tuck-in TEV multiple", 5.0, "TBD", "Placeholder only. Market notes suggest 7-10x tuck-ins, likely too high for G&B without a wedge."),
        ("Annual aquihire TEV", "=SUM(D10:I10)*$D$51*$D$32", "", ""),
        ("Annual tuck-in TEV", "=SUM(D11:I11)*$D$52*$D$33", "", ""),
        ("Total modeled acquisition TEV", "=D53+D54", "", "Use only as a directional capital need; no acquisition-by-acquisition timing yet."),
    ]
    for r, row in enumerate(cash_rows, start=51):
        ws.cell(r, 2).value = row[0]
        if isinstance(row[1], str) and row[1].startswith("="):
            formula(ws.cell(r, 4), row[1])
        else:
            input_cell(ws.cell(r, 4), row[1])
        ws.cell(r, 10).value = row[2]
        note(ws.cell(r, 11), row[3])

    money_fmt(ws, ["D18:I22", "D27:D34", "D42:I46", "D53:D55"])
    mult_fmt(ws, ["D51:D52"])
    style_sheet(ws)
    return ws


def build_stress_tests(wb):
    ws = wb.create_sheet("Stress Tests", 2)
    setup_widths(ws, {"A": 3, "B": 26, "C": 14, "D": 14, "E": 14, "F": 14, "G": 18, "H": 44})
    ws["B2"] = "Seller Note / Rollover / Interest-Rate Stress Tests"
    ws["B2"].font = Font(bold=True, size=16, color=BLUE)
    ws["B3"] = "All tests use the base-case adjusted EBITDA path and change structure levers only."

    headers = ["Scenario", "Seller note %", "Rate", "Tenor", "Rollover %", "Min DSCR Y1-Y5", "Verdict"]
    for c, h in enumerate(headers, start=2):
        header(ws.cell(6, c), h)
    scenarios = [
        ("Low seller note / low rate", 0.20, 0.06, 7, 0.10),
        ("Base case", 0.30, 0.07, 6, 0.125),
        ("High seller note / high rate", 0.40, 0.08, 5, 0.15),
        ("No rollover / more cash need", 0.30, 0.07, 6, 0.00),
        ("Max rollover / max seller carry", 0.40, 0.08, 7, 0.15),
        ("Interest shock", 0.30, 0.09, 6, 0.125),
    ]
    for r, (name, sn, rate, tenor, roll) in enumerate(scenarios, start=7):
        ws.cell(r, 2).value = name
        for c, val in zip(range(3, 7), [sn, rate, tenor, roll]):
            input_cell(ws.cell(r, c), val)
        denom = f"((\'Pest Base Case\'!$D$18*C{r}*D{r})+(\'Pest Base Case\'!$D$18*C{r}/E{r}))"
        formula(ws.cell(r, 7), f'=MIN(\'Pest Base Case\'!E64/{denom},\'Pest Base Case\'!F64/{denom},\'Pest Base Case\'!G64/{denom},\'Pest Base Case\'!H64/{denom},\'Pest Base Case\'!I64/{denom})')
        formula(ws.cell(r, 8), f'=IF(G{r}<1,"Breaks cash flow",IF(G{r}<1.25,"Marginal cushion","Supported in first pass"))')

    label(ws["B16"], "Cash-flow sensitivity")
    headers = ["EBITDA haircut", "Min DSCR base structure", "Cash after debt Y1", "Cash after debt Y5", "Verdict"]
    for c, h in enumerate(headers, start=2):
        header(ws.cell(18, c), h)
    for r, haircut in enumerate([0, 0.10, 0.20, 0.30], start=19):
        input_cell(ws.cell(r, 2), haircut)
        denom = "((\'Pest Base Case\'!$D$36*\'Pest Base Case\'!$D$23)+(\'Pest Base Case\'!$D$36/\'Pest Base Case\'!$D$24))"
        formula(ws.cell(r, 3), f'=MIN((\'Pest Base Case\'!E64*(1-B{r}))/{denom},(\'Pest Base Case\'!F64*(1-B{r}))/{denom},(\'Pest Base Case\'!G64*(1-B{r}))/{denom},(\'Pest Base Case\'!H64*(1-B{r}))/{denom},(\'Pest Base Case\'!I64*(1-B{r}))/{denom})')
        formula(ws.cell(r, 4), f'=\'Pest Base Case\'!E70*(1-B{r})')
        formula(ws.cell(r, 5), f'=\'Pest Base Case\'!I70*(1-B{r})')
        formula(ws.cell(r, 6), f'=IF(C{r}<1,"Breaks",IF(C{r}<1.25,"Tight","OK"))')

    pct_fmt(ws, ["C7:D12", "F7:F12", "B19:B22"])
    ws["E7:E12"][0][0].number_format = "0.0"
    for row in ws["E7:E12"]:
        row[0].number_format = "0.0"
    for row in ws["G7:G12"]:
        row[0].number_format = "0.00x"
    for row in ws["C19:C22"]:
        row[0].number_format = "0.00x"
    money_fmt(ws, ["D19:E22"])
    style_sheet(ws)
    return ws


def build_questions(wb):
    ws = wb.create_sheet("Investor Questions", 3)
    setup_widths(ws, {"A": 3, "B": 20, "C": 54, "D": 52, "E": 22})
    ws["B2"] = "Investor-Facing Question Set: Jeff / Anacapa"
    ws["B2"].font = Font(bold=True, size=16, color=BLUE)
    headers = ["Topic", "Question", "Why it matters", "Owner / next use"]
    for c, h in enumerate(headers, start=2):
        header(ws.cell(5, c), h)
    rows = [
        ("Cap table", "If the first pest platform is only ~$500K EBITDA, what cap-table shape would Anacapa support: same committed investor base, smaller first-close syndicate, or reserved follow-on capital?", "Melissa says committed capital should bridge the first close to the flywheel; Kay needs investor appetite before overbuilding the strategy.", "Jeff / Anacapa"),
        ("Seller note", "At what seller-note quantum does Anacapa become uncomfortable for a small, owner-operated pest platform?", "Seller note is the primary lever to make the seller's proceeds work without overloading cash flow.", "Jeff / Anacapa"),
        ("Rollover", "Is 10-15% seller rollover acceptable in a search-style cap table if the seller has no governance control?", "Jay suggested 10-15% max; Melissa sees low governance risk, but investor support matters.", "Jeff / Anacapa"),
        ("Interest rate", "Would Anacapa underwrite a 6-8% seller note today, and what cushion would they require if rates move up?", "Debt-service pressure is the gating issue for a small platform.", "Jeff / Anacapa"),
        ("Bank debt", "Does the ~$2M EBITDA / ~$5M facility threshold match Anacapa's current lender read for pest services?", "Model assumes no bank debt at close and future debt only after scale.", "Jeff / Anacapa"),
        ("Buy-box fit", "Is an undersized but high-margin $500K EBITDA pest platform acceptable if the multi-acquisition path is credible?", "CIO issue: initial asset is below preferred scale, but could be a wedge into a dense market.", "Jeff / Anacapa"),
        ("Wedge", "Does the luxury retail / hospitality discretion wedge make pest more underwritable, or is it too narrow for an investor-backed strategy?", "Jay's warning: broad pest is too competitive; specialty is the opening.", "Jeff / Anacapa"),
        ("Diligence", "What proof would make the acquisition path real: client count, route density, revenue per technician, retention, pricing uplift, or signed seller interest?", "Avoids modeling sophistication without diligence-grade inputs.", "Jeff / Anacapa"),
    ]
    for r, row in enumerate(rows, start=6):
        for c, val in enumerate(row, start=2):
            note(ws.cell(r, c), val)
    style_sheet(ws)
    return ws


def build_sources(wb):
    ws = wb.create_sheet("Source Notes", 4)
    setup_widths(ws, {"A": 3, "B": 24, "C": 24, "D": 72, "E": 36})
    ws["B2"] = "Source Notes / Assumption Register"
    ws["B2"].font = Font(bold=True, size=16, color=BLUE)
    headers = ["Source", "Model use", "Fact / guidance used", "Status"]
    for c, h in enumerate(headers, start=2):
        header(ws.cell(5, c), h)
    rows = [
        ("Melissa Rosenblatt call 2026-06-17", "Core model architecture", "Core business: $1.5M revenue, $500K EBITDA, 40% labor/COGS, no office manager, answering service. Build core IS, corporate cost section, M&A schedule, combined IS.", "Used"),
        ("Melissa analysis Google Doc 2026-06-17", "Operating frame", "Full analysis reinforced aquihire over de novo, holdco vs branded roll-up as open, and the need for one visible model plus a decision memo.", "Used"),
        ("Melissa transcript Google Doc 2026-06-17", "Transcript check", "gog docs cat returned only: Granola transcript unavailable in response. No transcript-only facts were used.", "Blocked / explicit limitation"),
        ("Melissa Rosenblatt call 2026-06-17", "Capital structure", "First acquisition likely equity + seller note; no bank debt at this size. Bank debt threshold ~$2M EBITDA, facility ~$5M. Seller note 5-7 years, 6-8% interest.", "Used"),
        ("Melissa Rosenblatt / Jay Davis 2026-06-17", "Rollover", "Jay suggested 10-15% max; Melissa confirmed minority stake carries no real governance risk.", "Used"),
        ("Guillermo Lavergne brainstorm 2026-06-17", "Target diligence", "NJ/NYC commercial pest company: ~$1.5M revenue, ~$500K EBITDA, 6-person team, 98% retention, Cartier/Chelsea Market/Google proof points.", "Used"),
        ("Carlos check-in 2026-06-17", "Thesis support", "Carlos validated small NJ operator / NYC thesis; holding-company structure remains live.", "Used"),
        ("WSN pest calibration 2026-06-17", "Diligence gate", "Cash-flow quality and owner draw need verification before more time is spent.", "Used as missing-input flag"),
        ("Jay Davis 2026-06-09", "Market structure", "Broad pest is crowded/liquid; specialty/premium wedge is the opening if growth is real.", "Used"),
        ("Premium Pest Scorecard Apr 2026", "Industry context", "19,000+ firms, strong fragmentation, 25-35% EBITDA margins for well-run operators, asset-light service model.", "Used as context, not hardcoded forecast"),
        ("Niche intelligence 2026-05-26", "Market multiple context", "PE consolidation deepening; 10-15x platform / 7-10x tuck-in market notes.", "Flagged: not used as base valuation because it may not fit G&B's small target"),
        ("Richard / Stony Hill 2026-06-17", "Comparable check", "Machinery and sports-surface listings are deal-flow checks, not pest comps. Machinery is capital-intensive/nonlocal; sports surfaces lacks enough relevance.", "Not used in model"),
        ("G&B Financial Model Template.xlsx", "Baseline structure", "Master template from Manager Documents / G&B Master Templates. Sources/uses, cash flow, debt schedule, returns framework retained and used as the reference single-acquisition structure.", "Used as baseline"),
    ]
    for r, row in enumerate(rows, start=6):
        for c, val in enumerate(row, start=2):
            note(ws.cell(r, c), val)
    style_sheet(ws)
    return ws


def main():
    wb = load_workbook(BASELINE)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    # Keep baseline tabs intact, but make their source status explicit.
    if "Model Template" in wb.sheetnames:
        wb["Model Template"]["B2"] = "G&B Financial Model Template - Baseline Reference"

    build_pest_single_acquisition(wb)
    build_base_case(wb)
    build_ma_schedule(wb)
    build_stress_tests(wb)
    build_questions(wb)
    build_sources(wb)

    preferred_order = [
        "Pest Single Acquisition",
        "Pest Base Case",
        "M&A Schedule",
        "Stress Tests",
        "Investor Questions",
        "Source Notes",
        "Model Template",
        "Growth",
        "Value",
    ]
    wb._sheets = [wb[name] for name in preferred_order if name in wb.sheetnames]

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
