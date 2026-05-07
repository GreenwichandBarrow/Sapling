"""Build Kay's personal task tracker Excel file.

Three tabs:
  - This Week:
      · Habit tracker block at top (label + Mon-Sun grid).
      · Day grid below: each day has date, donut progress, % done,
        5 priority slots, and a notes/journal area.
  - Backlog: rolling overflow list, items live here until done.
  - Projects: longer-than-a-day work with status + next action.

Aesthetic: sage green from the Instagram template. European Mon-Sun week.
No time-blocking. No macros. Strikethrough + fill when status flips to ✅ / Done.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties

import os
OUT = os.environ.get(
    "TASKS_XLSX_OUT",
    "/Users/kaycschneider/My Drive/STRATEGIC PLANNING/TO DO 4.26.26.xlsx",
)

SAGE_DARK = "6B8E5A"
SAGE_MID = "A8C49A"
SAGE_LIGHT = "E8F0E2"
CREAM = "FAF8F2"
INK = "2E3D2A"
MUTED = "8A8A7E"
RING_EMPTY = "E0E0D8"

HEADER_FONT = Font(name="Avenir Next", size=13, bold=True, color="FFFFFF")
DAY_FONT = Font(name="Avenir Next", size=12, bold=True, color="FFFFFF")
SUBHEAD_FONT = Font(name="Avenir Next", size=11, bold=True, color=INK)
BODY_FONT = Font(name="Avenir Next", size=10, color=INK)
HABIT_LABEL_FONT = Font(name="Avenir Next", size=10, bold=True, color=INK)
MUTED_FONT = Font(name="Avenir Next", size=9, italic=True, color=MUTED)
PCT_FONT = Font(name="Avenir Next", size=14, bold=True, color=SAGE_DARK)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin = Side(border_style="thin", color="D8DCD0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

DONE_FILL = PatternFill("solid", fgColor=SAGE_LIGHT)
HEADER_FILL = PatternFill("solid", fgColor=SAGE_DARK)
DAY_FILL = PatternFill("solid", fgColor=SAGE_MID)
SUBHEAD_FILL = PatternFill("solid", fgColor=SAGE_LIGHT)
ROW_FILL = PatternFill("solid", fgColor=CREAM)


def style_cell(cell, font, fill=None, align=CENTER, border=True):
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = align
    if border:
        cell.border = BORDER


def add_status_dropdown(ws, ranges):
    dv = DataValidation(type="list", formula1='"☐,✅"', allow_blank=True)
    for r in ranges:
        dv.add(r)
    ws.add_data_validation(dv)


# ---------- Workbook ----------
wb = Workbook()

# ============ THIS WEEK ============
week = wb.active
week.title = "This Week"
week.sheet_view.showGridLines = False
week.sheet_properties.tabColor = SAGE_DARK

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
DAY_ABBREV = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
# Each day spans 2 cols: small status (4 wide) + wide task (18 wide)
DAY_PAIRS = [("B","C"), ("D","E"), ("F","G"), ("H","I"), ("J","K"), ("L","M"), ("N","O")]

# Column widths
week.column_dimensions["A"].width = 22  # Habit labels / left margin
for sc, tc in DAY_PAIRS:
    week.column_dimensions[sc].width = 4
    week.column_dimensions[tc].width = 18

# Title banner row 2 (spans A:O)
week.merge_cells("A2:O2")
title = week["A2"]
title.value = "THIS WEEK"
style_cell(title, HEADER_FONT, HEADER_FILL, CENTER, border=False)
week.row_dimensions[2].height = 30

# Week-of row 3
week.merge_cells("A3:O3")
weekof = week["A3"]
weekof.value = '=" Week of "&TEXT(TODAY()-WEEKDAY(TODAY(),2)+1,"mmmm d, yyyy")'
weekof.font = Font(name="Avenir Next", size=11, italic=True, color=MUTED)
weekof.alignment = CENTER
weekof.fill = ROW_FILL
week.row_dimensions[3].height = 22

# ----- HABIT TRACKER block -----
# Row 5: section banner
week.merge_cells("A5:O5")
htitle = week["A5"]
htitle.value = "HABIT TRACKER"
style_cell(htitle, HEADER_FONT, HEADER_FILL, CENTER, border=False)
week.row_dimensions[5].height = 24

# Row 6: day abbreviations (each day's pair merged)
week["A6"] = ""
week["A6"].fill = ROW_FILL
for (sc, tc), abbrev in zip(DAY_PAIRS, DAY_ABBREV):
    week.merge_cells(f"{sc}6:{tc}6")
    cell = week[f"{sc}6"]
    cell.value = abbrev
    style_cell(cell, SUBHEAD_FONT, SUBHEAD_FILL, CENTER)
week.row_dimensions[6].height = 22

# Rows 7-13: 7 habit items, each day's checkbox cell merged across the pair
HABIT_ROWS = list(range(7, 14))
HABIT_SEED = [
    "Water & hygiene",
    "Meditation & stretches",
    "ACV drink & probiotic protein shake",
    "Exercise class",
    "Bike to work",
    "10K steps",
    "Omega 3 & magnesium",
]
for r, seed in zip(HABIT_ROWS, HABIT_SEED):
    week.row_dimensions[r].height = 24
    label = week[f"A{r}"]
    label.value = seed
    style_cell(label, HABIT_LABEL_FONT, ROW_FILL, LEFT)
    for sc, tc in DAY_PAIRS:
        week.merge_cells(f"{sc}{r}:{tc}{r}")
        cell = week[f"{sc}{r}"]
        cell.value = "☐"
        style_cell(cell, Font(name="Avenir Next", size=14, color=INK), ROW_FILL, CENTER)

# Strikethrough + sage fill on completed habit cells (status lives in the status col)
for sc, tc in DAY_PAIRS:
    rng = f"{sc}7:{sc}13"
    rule = FormulaRule(
        formula=[f'={sc}7="✅"'],
        font=Font(name="Avenir Next", size=14, color=SAGE_DARK, strike=False, bold=True),
        fill=DONE_FILL,
    )
    week.conditional_formatting.add(rng, rule)

add_status_dropdown(week, [f"{sc}7:{sc}13" for sc, tc in DAY_PAIRS])

# Spacer row 14
week.row_dimensions[14].height = 12

# ----- DAY GRID: starts row 15 -----
# Row 15: Day name headers (merged per day pair)
for (sc, tc), day in zip(DAY_PAIRS, DAYS):
    week.merge_cells(f"{sc}15:{tc}15")
    cell = week[f"{sc}15"]
    cell.value = day.upper()
    style_cell(cell, DAY_FONT, DAY_FILL, CENTER)
week["A15"].fill = ROW_FILL
week.row_dimensions[15].height = 22

# Row 16: Dates (merged per day pair)
for i, (sc, tc) in enumerate(DAY_PAIRS):
    week.merge_cells(f"{sc}16:{tc}16")
    cell = week[f"{sc}16"]
    cell.value = f'=TEXT(TODAY()-WEEKDAY(TODAY(),2)+1+{i},"mmm d")'
    style_cell(cell, BODY_FONT, ROW_FILL, CENTER)
week["A16"].fill = ROW_FILL
week.row_dimensions[16].height = 20

# Rows 17-21: Donut chart anchor area, merged per day as one bordered cell
DONUT_ROWS = list(range(17, 22))
donut_side = Side(border_style="thin", color=SAGE_MID)
donut_border = Border(left=donut_side, right=donut_side, top=donut_side, bottom=donut_side)
for r in DONUT_ROWS:
    week.row_dimensions[r].height = 18
    week[f"A{r}"].fill = ROW_FILL
for sc, tc in DAY_PAIRS:
    week.merge_cells(f"{sc}17:{tc}21")
    cell = week[f"{sc}17"]
    cell.fill = ROW_FILL
    cell.border = donut_border

# Row 22: % text below donut (merged per day pair)
week.row_dimensions[22].height = 22
week["A22"].fill = ROW_FILL
for sc, tc in DAY_PAIRS:
    week.merge_cells(f"{sc}22:{tc}22")
    pct = week[f"{sc}22"]
    pct.value = (
        f'=IF(COUNTA({tc}23:{tc}37)=0,"—",'
        f'TEXT(COUNTIF({sc}23:{sc}37,"✅")/COUNTA({tc}23:{tc}37),"0%"))'
    )
    style_cell(pct, PCT_FONT, ROW_FILL, CENTER)

# Rows 23-37: 15 priorities, status cell + task cell side by side per day
PRIORITY_ROWS = list(range(23, 38))
for r in PRIORITY_ROWS:
    week.row_dimensions[r].height = 28
    week[f"A{r}"].fill = ROW_FILL
    for sc, tc in DAY_PAIRS:
        scell = week[f"{sc}{r}"]
        scell.value = "☐"
        style_cell(scell, Font(name="Avenir Next", size=12, color=INK), ROW_FILL, CENTER)
        tcell = week[f"{tc}{r}"]
        tcell.value = ""
        style_cell(tcell, BODY_FONT, ROW_FILL, LEFT)

# Status dropdown on priority status cells
add_status_dropdown(week, [f"{sc}23:{sc}37" for sc, tc in DAY_PAIRS])

# Conditional format priorities: when status = ✅, strike+dim status+task pair
for sc, tc in DAY_PAIRS:
    rng = f"{sc}23:{tc}37"
    rule = FormulaRule(
        formula=[f'=${sc}23="✅"'],
        font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
        fill=DONE_FILL,
    )
    week.conditional_formatting.add(rng, rule)

# Spacer row 38
week.row_dimensions[38].height = 12

# Notes label row 39 (merged per day)
for sc, tc in DAY_PAIRS:
    week.merge_cells(f"{sc}39:{tc}39")
    cell = week[f"{sc}39"]
    cell.value = "notes"
    style_cell(cell, MUTED_FONT, SUBHEAD_FILL, CENTER)
week["A39"].fill = ROW_FILL
week.row_dimensions[39].height = 18

# Notes area rows 40-47 (merged per day) — slimmed from 24 rows to 8
for r in range(40, 48):
    week.row_dimensions[r].height = 16
    week[f"A{r}"].fill = ROW_FILL
    for sc, tc in DAY_PAIRS:
        week.merge_cells(f"{sc}{r}:{tc}{r}")
        cell = week[f"{sc}{r}"]
        cell.value = ""
        style_cell(cell, BODY_FONT, ROW_FILL, LEFT_TOP)

# ----- Helper data for donut charts (rows 99-107, hidden) -----
week["Q99"] = "Day"
week["R99"] = "Done"
week["S99"] = "Open"
for cell_ref in ["Q99", "R99", "S99"]:
    week[cell_ref].font = MUTED_FONT

for i, (sc, tc) in enumerate(DAY_PAIRS):
    r = 100 + i
    week[f"Q{r}"] = DAYS[i]
    week[f"Q{r}"].font = MUTED_FONT
    week[f"R{r}"] = f'=COUNTIF({sc}23:{sc}37,"✅")'
    week[f"S{r}"] = (
        f'=IF(COUNTA({tc}23:{tc}37)=0,1,'
        f'COUNTA({tc}23:{tc}37)-COUNTIF({sc}23:{sc}37,"✅"))'
    )

for r in range(99, 108):
    week.row_dimensions[r].hidden = True

# ----- Donut chart per day, anchored at top of donut area (row 17) -----
for i, (sc, tc) in enumerate(DAY_PAIRS):
    helper_row = 100 + i
    chart = DoughnutChart()
    chart.title = None
    chart.legend = None
    chart.holeSize = 65

    data_ref = Reference(
        week,
        min_col=18,  # R (Done helper)
        max_col=19,  # S (Open helper)
        min_row=helper_row,
        max_row=helper_row,
    )
    chart.add_data(data_ref, titles_from_data=False, from_rows=True)

    series = chart.series[0]
    done_pt = DataPoint(idx=0)
    done_pt.graphicalProperties = GraphicalProperties(solidFill=SAGE_DARK)
    open_pt = DataPoint(idx=1)
    open_pt.graphicalProperties = GraphicalProperties(solidFill=RING_EMPTY)
    series.data_points = [done_pt, open_pt]

    chart.width = 3.8
    chart.height = 3.5

    week.add_chart(chart, f"{sc}17")

# Freeze pane below habit tracker so day grid scrolls but tracker stays visible
week.freeze_panes = "A15"


# ============ TO DO (single capture point, was Backlog) ============
todo = wb.create_sheet("To Do")
todo.sheet_view.showGridLines = False
todo.sheet_properties.tabColor = SAGE_MID

twidths = {"A": 2, "B": 5, "C": 44, "D": 10, "E": 22, "F": 12, "G": 24}
for col, w in twidths.items():
    todo.column_dimensions[col].width = w

todo.merge_cells("B2:G2")
t_title = todo["B2"]
t_title.value = "TO DO  ·  capture everything here"
style_cell(t_title, HEADER_FONT, HEADER_FILL, CENTER, border=False)
todo.row_dimensions[2].height = 30

todo.merge_cells("B3:G3")
t_sub = todo["B3"]
t_sub.value = "Tag with Type + Project. Pull into This Week when ready."
t_sub.font = Font(name="Avenir Next", size=10, italic=True, color=MUTED)
t_sub.alignment = CENTER
t_sub.fill = ROW_FILL
todo.row_dimensions[3].height = 22

theaders = [
    ("B5", ""),
    ("C5", "Task"),
    ("D5", "Type"),
    ("E5", "Project"),
    ("F5", "Due"),
    ("G5", "Notes"),
]
for cell_ref, text in theaders:
    todo[cell_ref] = text
    style_cell(todo[cell_ref], SUBHEAD_FONT, DAY_FILL, CENTER)
todo.row_dimensions[5].height = 22

for row in range(6, 86):
    todo[f"B{row}"] = "☐"
    style_cell(todo[f"B{row}"], Font(name="Avenir Next", size=12, color=INK), ROW_FILL, CENTER)
    for col in ["C", "D", "E", "F", "G"]:
        cell = todo[f"{col}{row}"]
        cell.value = ""
        style_cell(cell, BODY_FONT, ROW_FILL, LEFT if col != "D" else CENTER)
    todo.row_dimensions[row].height = 22

add_status_dropdown(todo, ["B6:B85"])

# Type dropdown: Home / Work
type_dv = DataValidation(type="list", formula1='"Work,Home"', allow_blank=True)
type_dv.add("D6:D85")
todo.add_data_validation(type_dv)

todo_done = FormulaRule(
    formula=['=$B6="✅"'],
    font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
    fill=DONE_FILL,
)
todo.conditional_formatting.add("B6:G85", todo_done)

# Subtle Type tint
home_rule = FormulaRule(
    formula=['=$D6="Home"'],
    fill=PatternFill("solid", fgColor="F4E8D8"),
)
work_rule = FormulaRule(
    formula=['=$D6="Work"'],
    fill=PatternFill("solid", fgColor=SAGE_LIGHT),
)
todo.conditional_formatting.add("D6:D85", home_rule)
todo.conditional_formatting.add("D6:D85", work_rule)

todo.freeze_panes = "B6"


# ============ TO DO LONG TERM (was Projects, now intents without hard timelines) ============
longterm = wb.create_sheet("To Do Long Term")
longterm.sheet_view.showGridLines = False
longterm.sheet_properties.tabColor = "D4A574"

ltwidths = {"A": 2, "B": 5, "C": 38, "D": 10, "E": 14, "F": 38}
for col, w in ltwidths.items():
    longterm.column_dimensions[col].width = w

longterm.merge_cells("B2:F2")
lt_title = longterm["B2"]
lt_title.value = "TO DO LONG TERM  ·  no hard timeline"
style_cell(lt_title, HEADER_FONT, HEADER_FILL, CENTER, border=False)
longterm.row_dimensions[2].height = 30

longterm.merge_cells("B3:F3")
lt_sub = longterm["B3"]
lt_sub.value = "Intents and someday items. Promote to Projects tab when ready to plan."
lt_sub.font = Font(name="Avenir Next", size=10, italic=True, color=MUTED)
lt_sub.alignment = CENTER
lt_sub.fill = ROW_FILL
longterm.row_dimensions[3].height = 22

ltheaders = [
    ("B5", ""),
    ("C5", "Item"),
    ("D5", "Type"),
    ("E5", "Status"),
    ("F5", "Notes"),
]
for cell_ref, text in ltheaders:
    longterm[cell_ref] = text
    style_cell(longterm[cell_ref], SUBHEAD_FONT, DAY_FILL, CENTER)
longterm.row_dimensions[5].height = 22

for row in range(6, 41):
    longterm[f"B{row}"] = "☐"
    style_cell(longterm[f"B{row}"], Font(name="Avenir Next", size=12, color=INK), ROW_FILL, CENTER)
    for col in ["C", "D", "E", "F"]:
        cell = longterm[f"{col}{row}"]
        cell.value = ""
        align = CENTER if col in ("D", "E") else LEFT
        style_cell(cell, BODY_FONT, ROW_FILL, align)
    longterm.row_dimensions[row].height = 26

add_status_dropdown(longterm, ["B6:B40"])

lt_type_dv = DataValidation(type="list", formula1='"Work,Home"', allow_blank=True)
lt_type_dv.add("D6:D40")
longterm.add_data_validation(lt_type_dv)

lt_status_dv = DataValidation(
    type="list",
    formula1='"Idea,Active,On hold,Promoted,Done"',
    allow_blank=True,
)
lt_status_dv.add("E6:E40")
longterm.add_data_validation(lt_status_dv)

lt_done = FormulaRule(
    formula=['=$B6="✅"'],
    font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
    fill=DONE_FILL,
)
longterm.conditional_formatting.add("B6:F40", lt_done)

lt_home_rule = FormulaRule(formula=['=$D6="Home"'], fill=PatternFill("solid", fgColor="F4E8D8"))
lt_work_rule = FormulaRule(formula=['=$D6="Work"'], fill=PatternFill("solid", fgColor=SAGE_LIGHT))
longterm.conditional_formatting.add("D6:D40", lt_home_rule)
longterm.conditional_formatting.add("D6:D40", lt_work_rule)

longterm.freeze_panes = "B6"


# ============ PROJECTS (Gantt index — links to per-project tabs) ============
PROJECTS_TAB = wb.create_sheet("Projects")
PROJECTS_TAB.sheet_view.showGridLines = False
PROJECTS_TAB.sheet_properties.tabColor = SAGE_DARK

prwidths = {"A": 2, "B": 32, "C": 14, "D": 12, "E": 12, "F": 12, "G": 12, "H": 32}
for col, w in prwidths.items():
    PROJECTS_TAB.column_dimensions[col].width = w

PROJECTS_TAB.merge_cells("B2:H2")
pr_title = PROJECTS_TAB["B2"]
pr_title.value = "PROJECTS  ·  active time-bound projects"
style_cell(pr_title, HEADER_FONT, HEADER_FILL, CENTER, border=False)
PROJECTS_TAB.row_dimensions[2].height = 30

PROJECTS_TAB.merge_cells("B3:H3")
pr_sub = PROJECTS_TAB["B3"]
pr_sub.value = "Each row links to its own Gantt tab. Click Open to jump in."
pr_sub.font = Font(name="Avenir Next", size=10, italic=True, color=MUTED)
pr_sub.alignment = CENTER
pr_sub.fill = ROW_FILL
PROJECTS_TAB.row_dimensions[3].height = 22

prheaders = [
    ("B5", "Project"),
    ("C5", "Entity"),
    ("D5", "Status"),
    ("E5", "Start"),
    ("F5", "Target"),
    ("G5", "Tab"),
    ("H5", "Notes"),
]
for cell_ref, text in prheaders:
    PROJECTS_TAB[cell_ref] = text
    style_cell(PROJECTS_TAB[cell_ref], SUBHEAD_FONT, DAY_FILL, CENTER)
PROJECTS_TAB.row_dimensions[5].height = 22

for row in range(6, 26):
    for col in ["B", "C", "D", "E", "F", "G", "H"]:
        cell = PROJECTS_TAB[f"{col}{row}"]
        cell.value = ""
        align = CENTER if col in ("C", "D", "E", "F", "G") else LEFT
        style_cell(cell, BODY_FONT, ROW_FILL, align)
    PROJECTS_TAB.row_dimensions[row].height = 26

ENTITY_DROPDOWN = '"Home,G&amp;B,Myself Renewed,Kai Grey,Panthera Grey"'
pr_entity_dv = DataValidation(type="list", formula1=ENTITY_DROPDOWN, allow_blank=True)
pr_entity_dv.add("C6:C25")
PROJECTS_TAB.add_data_validation(pr_entity_dv)

pr_status_dv = DataValidation(
    type="list",
    formula1='"Planning,Active,On hold,Done"',
    allow_blank=True,
)
pr_status_dv.add("D6:D25")
PROJECTS_TAB.add_data_validation(pr_status_dv)

pr_done = FormulaRule(
    formula=['=$D6="Done"'],
    font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
    fill=DONE_FILL,
)
PROJECTS_TAB.conditional_formatting.add("B6:H25", pr_done)

# Per-entity color tint on the Entity cell
ENTITY_COLORS = {
    "Home": "F4E8D8",
    "G&B": "E8F0E2",
    "Myself Renewed": "F5E1E1",
    "Kai Grey": "E1ECF5",
    "Panthera Grey": "ECE1F5",
}
for label, color in ENTITY_COLORS.items():
    rule = FormulaRule(
        formula=[f'=$C6="{label}"'],
        fill=PatternFill("solid", fgColor=color),
    )
    PROJECTS_TAB.conditional_formatting.add("C6:C25", rule)

PROJECTS_TAB.freeze_panes = "B6"


# ============ MYSELF RENEWED HEALTHCARE (Gantt project tab) ============
def build_gantt_project_tab(workbook, project_name, entity, milestones, weeks_span=16):
    """Create a per-project Gantt tab.

    milestones: list of (status, name, start_date_str, target_date_str, notes)
        where dates are ISO 'YYYY-MM-DD'.
    weeks_span: number of week columns in the timeline.
    """
    from datetime import date, timedelta

    entity_color = {
        "Home": "F4E8D8",
        "G&B": "E8F0E2",
        "Myself Renewed": "F5E1E1",
        "Kai Grey": "E1ECF5",
        "Panthera Grey": "ECE1F5",
    }.get(entity, "E0E0D8")

    ws = workbook.create_sheet(project_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = entity_color

    # Column widths: A spacer, B-F milestone table, G+ weekly timeline
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 5      # status checkbox
    ws.column_dimensions["C"].width = 38     # milestone name
    ws.column_dimensions["D"].width = 12     # start
    ws.column_dimensions["E"].width = 12     # target
    ws.column_dimensions["F"].width = 22     # notes

    # Title banner
    last_col_idx = 6 + weeks_span  # F is col 6; weeks start at G (col 7)
    last_col_letter = chr(ord("A") + last_col_idx - 1) if last_col_idx <= 26 else None
    # Handle >26 columns
    from openpyxl.utils import get_column_letter
    last_col_letter = get_column_letter(last_col_idx)

    ws.merge_cells(f"B2:{last_col_letter}2")
    title = ws["B2"]
    title.value = project_name.upper()
    style_cell(title, HEADER_FONT, HEADER_FILL, CENTER, border=False)
    ws.row_dimensions[2].height = 30

    ws.merge_cells(f"B3:{last_col_letter}3")
    sub = ws["B3"]
    sub.value = f"Entity: {entity}  ·  Tick the week boxes you're actively working on each milestone — the row builds into a Gantt bar"
    sub.font = Font(name="Avenir Next", size=10, italic=True, color=MUTED)
    sub.alignment = CENTER
    sub.fill = ROW_FILL
    ws.row_dimensions[3].height = 22

    # Milestone table headers (row 5)
    headers = [("B5", ""), ("C5", "Milestone"), ("D5", "Start"), ("E5", "Target"), ("F5", "Notes")]
    for cell_ref, text in headers:
        ws[cell_ref] = text
        style_cell(ws[cell_ref], SUBHEAD_FONT, DAY_FILL, CENTER)

    # Week column headers (row 5)
    today = date.today()
    monday = today - timedelta(days=today.weekday())
    for w in range(weeks_span):
        col_idx = 7 + w
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 5
        week_start = monday + timedelta(weeks=w)
        cell = ws[f"{col_letter}5"]
        cell.value = week_start.strftime("%-m/%-d")
        style_cell(cell, MUTED_FONT, DAY_FILL, CENTER)

    ws.row_dimensions[5].height = 22

    # Milestone rows starting at row 6
    for i, (status, name, start, target, notes) in enumerate(milestones):
        row = 6 + i
        ws.row_dimensions[row].height = 24

        ws[f"B{row}"] = status or "☐"
        style_cell(ws[f"B{row}"], Font(name="Avenir Next", size=12, color=INK), ROW_FILL, CENTER)

        ws[f"C{row}"] = name
        style_cell(ws[f"C{row}"], BODY_FONT, ROW_FILL, LEFT)

        if start:
            ws[f"D{row}"] = start
            ws[f"D{row}"].number_format = "yyyy-mm-dd"
        style_cell(ws[f"D{row}"], BODY_FONT, ROW_FILL, CENTER)

        if target:
            ws[f"E{row}"] = target
            ws[f"E{row}"].number_format = "yyyy-mm-dd"
        style_cell(ws[f"E{row}"], BODY_FONT, ROW_FILL, CENTER)

        ws[f"F{row}"] = notes or ""
        style_cell(ws[f"F{row}"], BODY_FONT, ROW_FILL, LEFT)

        # Timeline cells: each is a manual checkbox. Tick to fill the cell
        # entity-color so the row builds into a Gantt bar visually.
        for w in range(weeks_span):
            col_idx = 7 + w
            col_letter = get_column_letter(col_idx)
            tl = ws[f"{col_letter}{row}"]
            tl.value = "☐"
            tl.font = Font(name="Avenir Next", size=10, color=INK)
            tl.alignment = CENTER
            tl.fill = ROW_FILL
            tl.border = Border(
                left=Side(border_style="thin", color="E5E5DD"),
                right=Side(border_style="thin", color="E5E5DD"),
            )

    last_milestone_row = 5 + len(milestones)

    # Status dropdown on milestone status column
    add_status_dropdown(ws, [f"B6:B{last_milestone_row}"])

    # Status dropdown on every timeline cell
    add_status_dropdown(ws, [f"G6:{last_col_letter}{last_milestone_row}"])

    # Per-cell rule: ✅ in a timeline cell fills it with entity color
    timeline_range = f"G6:{last_col_letter}{last_milestone_row}"
    bar_rule = FormulaRule(
        formula=['=G6="✅"'],
        fill=PatternFill("solid", fgColor=entity_color),
    )
    ws.conditional_formatting.add(timeline_range, bar_rule)

    # Done strikethrough on the milestone left-side table
    done_rule = FormulaRule(
        formula=['=$B6="✅"'],
        font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
        fill=DONE_FILL,
    )
    ws.conditional_formatting.add(f"B6:F{last_milestone_row}", done_rule)

    ws.freeze_panes = "G6"


HEALTHCARE_MILESTONES = [
    ("☐", "Verify 501(c)(3) status active with NY Dept of State", "2026-04-27", "2026-05-08", ""),
    ("☐", "Confirm registration with NY Charities Bureau", "2026-04-27", "2026-05-08", ""),
    ("☐", "Hold board meeting + adopt written resolution", "2026-05-11", "2026-05-22", ""),
    ("☐", "Confirm nonprofit bank account + PayPal Giving Fund integration", "2026-05-18", "2026-06-05", ""),
    ("☐", "Add Myself Renewed to QuickBooks (nonprofit category)", "2026-05-25", "2026-06-05", ""),
    ("☐", "Set up donation processor + add donation page language", "2026-06-01", "2026-06-19", ""),
    ("☐", "Verify recipient 501(c)(3); board vote + grant approval", "2026-06-22", "2026-07-10", ""),
    ("☐", "Make first personal donation of $1,000", "2026-06-29", "2026-07-03", ""),
    ("☐", "Draft + publish personal giving statement aligned with legacy vision", "2026-06-01", "2026-07-17", ""),
    ("☐", "Healthcare program design — define plan tiers + delivery model", "2026-07-06", "2026-08-14", "Kay to detail"),
]
build_gantt_project_tab(wb, "Myself Renewed Healthcare", "Myself Renewed", HEALTHCARE_MILESTONES, weeks_span=16)


wb.save(OUT)
print(f"Wrote {OUT}")
