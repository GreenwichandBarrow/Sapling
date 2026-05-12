#!/usr/bin/env python3
"""One-shot Sheet builder for TO DO 5.12.26 — Phase 1 (structure) + Phase 2 (data migration).

Creates a new Google Sheet in the STRATEGIC PLANNING Drive folder with the same
5-tab architecture as the source Excel, applies native Sheets checkboxes /
dropdowns / conditional formatting, migrates all data from the source Excel,
and adds today's Saltoun To Do row + promotes it to today's slot 1.

Run once:
    /home/ubuntu/projects/Sapling/dashboard/.venv/bin/python3 build_sheet.py

Prints the new Sheet ID + URL on success.
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
import tempfile
import time
from datetime import date, timedelta
from pathlib import Path

import requests
from openpyxl import load_workbook

# ---------------------------------------------------------------- config

PARENT_FOLDER_ID = "12IpnsQ5V_M1fiTm0NZM9wKhlerauILMd"  # STRATEGIC PLANNING
SOURCE_XLSX_ID = "1rMhPMuG6I-NElbocYQE3LQD4Inlzg6on"
SOURCE_XLSX_PATH = "/tmp/tracker-migration/source.xlsx"
NEW_SHEET_TITLE = "TO DO 5.12.26"
GOG_ACCOUNT = os.environ.get("GOG_ACCOUNT", "kay.s@greenwichandbarrow.com")

GOG_CREDS_PATH = Path.home() / ".config" / "gogcli" / "credentials.json"

# Palette (RGB 0..1 for Sheets API)
def hex_to_rgb(hexstr: str) -> dict:
    h = hexstr.lstrip("#")
    return {
        "red": int(h[0:2], 16) / 255.0,
        "green": int(h[2:4], 16) / 255.0,
        "blue": int(h[4:6], 16) / 255.0,
    }


SAGE_LIGHT_HEX = "e8efd8"  # background
SAGE_DARK_HEX = "7a8c4d"  # accent
SAGE_EXTRA_LIGHT_HEX = "f3f7e8"  # done fill
INK_HEX = "2e3d2a"
MUTED_HEX = "9a9a8a"

# Entity colors for Gantt fills
ENTITY_COLOR_HEX = {
    "G&B": "7a8c4d",
    "Kai Grey": "9b8e7c",
    "Panthera Grey": "7a7e89",
    "Myself Renewed": "f4ddd9",
    "Home": "d8c7a8",
}

# Type colors for To Do / To Do Long Term Type column
TYPE_HOME_HEX = "f4e8d8"
TYPE_WORK_HEX = SAGE_EXTRA_LIGHT_HEX

# Day-of-week labels for the live week tab
WEEKDAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

# Habit list (read from source Excel rows 7-13 col A)
HABITS = [
    "Water & hygiene",
    "Meditation & stretches",
    "ACV drink & probiotic protein shake",
    "Exercise class",
    "Bike to work",
    "10K steps",
    "Omega 3 & magnesium",
]


# ---------------------------------------------------------------- auth

def get_access_token() -> str:
    creds = json.loads(GOG_CREDS_PATH.read_text())
    with tempfile.NamedTemporaryFile("r", suffix=".json", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    try:
        export = subprocess.run(
            ["gog", "auth", "tokens", "export", GOG_ACCOUNT, "--out", str(tmp_path), "--overwrite"],
            capture_output=True, text=True, timeout=15,
        )
        if export.returncode != 0:
            raise RuntimeError(f"gog token export failed: {export.stderr[:200]}")
        token_file = json.loads(tmp_path.read_text())
        refresh_token = token_file["refresh_token"]
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except OSError:
            pass

    resp = requests.post(
        "https://oauth2.googleapis.com/token",
        data={
            "client_id": creds["client_id"],
            "client_secret": creds["client_secret"],
            "refresh_token": refresh_token,
            "grant_type": "refresh_token",
        },
        timeout=15,
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


# ---------------------------------------------------------------- API helpers

class SheetsClient:
    def __init__(self, token: str):
        self.token = token
        self.session = requests.Session()
        self.session.headers.update({"Authorization": f"Bearer {token}"})

    def create_spreadsheet(self, title: str) -> dict:
        body = {"properties": {"title": title}}
        r = self.session.post("https://sheets.googleapis.com/v4/spreadsheets", json=body, timeout=30)
        r.raise_for_status()
        return r.json()

    def batch_update(self, spreadsheet_id: str, requests_list: list) -> dict:
        body = {"requests": requests_list}
        for attempt in range(5):
            r = self.session.post(
                f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}:batchUpdate",
                json=body, timeout=60,
            )
            if r.status_code == 200:
                return r.json()
            if r.status_code == 429 or r.status_code >= 500:
                wait = 2 ** attempt
                print(f"[backoff] HTTP {r.status_code} retrying in {wait}s...", file=sys.stderr)
                time.sleep(wait)
                continue
            # Surface error detail for debugging
            try:
                err = r.json()
                print(f"[ERROR] HTTP {r.status_code}: {json.dumps(err, indent=2)[:2000]}", file=sys.stderr)
            except Exception:
                print(f"[ERROR] HTTP {r.status_code}: {r.text[:2000]}", file=sys.stderr)
            r.raise_for_status()
        r.raise_for_status()

    def values_update(self, spreadsheet_id: str, range_a1: str, values: list[list]) -> dict:
        for attempt in range(5):
            r = self.session.put(
                f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{range_a1}",
                params={"valueInputOption": "USER_ENTERED"},
                json={"values": values},
                timeout=60,
            )
            if r.status_code == 200:
                return r.json()
            if r.status_code == 429 or r.status_code >= 500:
                time.sleep(2 ** attempt)
                continue
            r.raise_for_status()
        r.raise_for_status()


def move_to_folder(token: str, file_id: str, parent_id: str) -> None:
    # Find current parent
    r = requests.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        headers={"Authorization": f"Bearer {token}"},
        params={"fields": "parents"},
        timeout=30,
    )
    r.raise_for_status()
    old_parents = ",".join(r.json().get("parents", []))
    r = requests.patch(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        headers={"Authorization": f"Bearer {token}"},
        params={"addParents": parent_id, "removeParents": old_parents},
        timeout=30,
    )
    r.raise_for_status()


# ---------------------------------------------------------------- Sheet builders

def build_grid_properties(sheet_id: int, title: str, rows: int, cols: int,
                          frozen_rows: int = 0, frozen_cols: int = 0,
                          tab_color_hex: str | None = None,
                          index: int | None = None) -> dict:
    p = {
        "sheetId": sheet_id,
        "title": title,
        "gridProperties": {
            "rowCount": rows,
            "columnCount": cols,
            "frozenRowCount": frozen_rows,
            "frozenColumnCount": frozen_cols,
        },
    }
    if tab_color_hex:
        p["tabColor"] = hex_to_rgb(tab_color_hex)
    if index is not None:
        p["index"] = index
    return p


def add_sheet_request(props: dict) -> dict:
    return {"addSheet": {"properties": props}}


def update_dim_request(sheet_id: int, dim: str, start: int, end: int, pixel: int) -> dict:
    """dim is ROWS or COLUMNS, indices are 0-based half-open."""
    return {
        "updateDimensionProperties": {
            "range": {
                "sheetId": sheet_id,
                "dimension": dim,
                "startIndex": start,
                "endIndex": end,
            },
            "properties": {"pixelSize": pixel},
            "fields": "pixelSize",
        }
    }


def merge_cells_request(sheet_id: int, start_row: int, end_row: int,
                       start_col: int, end_col: int) -> dict:
    return {
        "mergeCells": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": start_row,
                "endRowIndex": end_row,
                "startColumnIndex": start_col,
                "endColumnIndex": end_col,
            },
            "mergeType": "MERGE_ALL",
        }
    }


def format_range_request(sheet_id: int, start_row: int, end_row: int,
                         start_col: int, end_col: int, fmt: dict, fields: str) -> dict:
    return {
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": start_row,
                "endRowIndex": end_row,
                "startColumnIndex": start_col,
                "endColumnIndex": end_col,
            },
            "cell": {"userEnteredFormat": fmt},
            "fields": fields,
        }
    }


def set_value_request(sheet_id: int, row: int, col: int, value, fmt: dict | None = None,
                      formula: bool = False) -> dict:
    if formula and isinstance(value, str) and value.startswith("="):
        ue = {"formulaValue": value}
    elif isinstance(value, bool):
        ue = {"boolValue": value}
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        ue = {"numberValue": value}
    elif value is None:
        ue = {}
    else:
        ue = {"stringValue": str(value)}
    cell = {"userEnteredValue": ue}
    fields = "userEnteredValue"
    if fmt:
        cell["userEnteredFormat"] = fmt
        fields += ",userEnteredFormat"
    return {
        "updateCells": {
            "rows": [{"values": [cell]}],
            "fields": fields,
            "start": {
                "sheetId": sheet_id,
                "rowIndex": row,
                "columnIndex": col,
            },
        }
    }


def checkbox_validation_request(sheet_id: int, start_row: int, end_row: int,
                                start_col: int, end_col: int) -> dict:
    return {
        "setDataValidation": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": start_row,
                "endRowIndex": end_row,
                "startColumnIndex": start_col,
                "endColumnIndex": end_col,
            },
            "rule": {
                "condition": {"type": "BOOLEAN"},
                "strict": True,
            },
        }
    }


def dropdown_validation_request(sheet_id: int, start_row: int, end_row: int,
                                start_col: int, end_col: int,
                                values: list[str], strict: bool = False) -> dict:
    return {
        "setDataValidation": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": start_row,
                "endRowIndex": end_row,
                "startColumnIndex": start_col,
                "endColumnIndex": end_col,
            },
            "rule": {
                "condition": {
                    "type": "ONE_OF_LIST",
                    "values": [{"userEnteredValue": v} for v in values],
                },
                "showCustomUi": True,
                "strict": strict,
            },
        }
    }


def conditional_format_request(sheet_id: int, start_row: int, end_row: int,
                                start_col: int, end_col: int,
                                formula: str, bg_hex: str | None = None,
                                text_hex: str | None = None,
                                strikethrough: bool = False,
                                bold: bool = False,
                                italic: bool = False,
                                index: int = 0) -> dict:
    fmt = {}
    if bg_hex:
        fmt["backgroundColor"] = hex_to_rgb(bg_hex)
    text_fmt = {}
    if text_hex:
        text_fmt["foregroundColor"] = hex_to_rgb(text_hex)
    if strikethrough:
        text_fmt["strikethrough"] = True
    if bold:
        text_fmt["bold"] = True
    if italic:
        text_fmt["italic"] = True
    if text_fmt:
        fmt["textFormat"] = text_fmt
    return {
        "addConditionalFormatRule": {
            "rule": {
                "ranges": [{
                    "sheetId": sheet_id,
                    "startRowIndex": start_row,
                    "endRowIndex": end_row,
                    "startColumnIndex": start_col,
                    "endColumnIndex": end_col,
                }],
                "booleanRule": {
                    "condition": {
                        "type": "CUSTOM_FORMULA",
                        "values": [{"userEnteredValue": formula}],
                    },
                    "format": fmt,
                },
            },
            "index": index,
        }
    }


# ---------------------------------------------------------------- Live Week tab

def col_letter(idx_0: int) -> str:
    """0-based index → A1 letter."""
    s = ""
    n = idx_0
    while True:
        s = chr(65 + n % 26) + s
        n = n // 26 - 1
        if n < 0:
            break
    return s


def live_week_label(today: date) -> str:
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)
    if monday.month == sunday.month:
        return f"{monday.strftime('%b')} {monday.day}-{sunday.day}"
    return f"{monday.strftime('%b')} {monday.day}-{sunday.strftime('%b')} {sunday.day}"


# Live Week layout uses 15 columns: A (label) + 7 day-pairs (status, task)
# A=0, Mon-stat=1, Mon-task=2, Tue-stat=3, ..., Sun-task=14
LIVE_DAY_STAT = {i: 1 + i * 2 for i in range(7)}
LIVE_DAY_TASK = {i: 2 + i * 2 for i in range(7)}


def build_live_week_requests(sheet_id: int, week_label: str) -> tuple[list[dict], list[dict]]:
    """Return (structure_requests, value_requests). Structure runs first; values after."""
    R: list[dict] = []
    V: list[dict] = []

    # Column widths
    R.append(update_dim_request(sheet_id, "COLUMNS", 0, 1, 180))  # A (habit label)
    # Each day-pair: status (narrow) + task (wide)
    for i in range(7):
        R.append(update_dim_request(sheet_id, "COLUMNS", LIVE_DAY_STAT[i], LIVE_DAY_STAT[i] + 1, 36))
        R.append(update_dim_request(sheet_id, "COLUMNS", LIVE_DAY_TASK[i], LIVE_DAY_TASK[i] + 1, 200))

    # Row heights
    R.append(update_dim_request(sheet_id, "ROWS", 0, 1, 42))  # title
    R.append(update_dim_request(sheet_id, "ROWS", 14, 15, 26))  # day header
    R.append(update_dim_request(sheet_id, "ROWS", 16, 21, 56))  # big % rows

    # Title row 1 (0-based row 0), merged A:O (cols 0..14 inclusive → 0..15 half-open)
    R.append(merge_cells_request(sheet_id, 0, 1, 0, 15))  # 15 cols total: 0..14

    V.append(set_value_request(sheet_id, 0, 0, f"KAY — WEEK OF {week_label}", fmt={
        "backgroundColor": hex_to_rgb(SAGE_LIGHT_HEX),
        "horizontalAlignment": "CENTER",
        "verticalAlignment": "MIDDLE",
        "textFormat": {"bold": True, "fontSize": 16, "foregroundColor": hex_to_rgb(SAGE_DARK_HEX)},
    }, formula=False))

    # HABIT TRACKER label row 5 (0-based row 4) col A
    V.append(set_value_request(sheet_id, 4, 0, "HABIT TRACKER", fmt={
        "textFormat": {"bold": True, "fontSize": 10, "foregroundColor": hex_to_rgb(SAGE_DARK_HEX)},
    }, formula=False))

    # Row 6 (0-based 5): day-of-week short labels above habit grid
    short_days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for i, d in enumerate(short_days):
        # Span the day-pair
        R.append(merge_cells_request(sheet_id, 5, 6, LIVE_DAY_STAT[i], LIVE_DAY_TASK[i] + 1))
        V.append(set_value_request(sheet_id, 5, LIVE_DAY_STAT[i], d, fmt={
            "horizontalAlignment": "CENTER",
            "textFormat": {"bold": True, "fontSize": 9, "foregroundColor": hex_to_rgb(MUTED_HEX)},
        }, formula=False))

    # Habit rows 7-13 (0-based 6..12): habit name in A, checkbox in each day-stat col
    for h_idx, habit in enumerate(HABITS):
        row = 6 + h_idx
        V.append(set_value_request(sheet_id, row, 0, habit, fmt={
            "textFormat": {"fontSize": 10, "foregroundColor": hex_to_rgb(INK_HEX)},
        }, formula=False))
        for i in range(7):
            R.append(checkbox_validation_request(sheet_id, row, row + 1, LIVE_DAY_STAT[i], LIVE_DAY_STAT[i] + 1))

    # Habit completion conditional formatting: if status TRUE → sage-light fill across day-pair
    for i in range(7):
        sc = LIVE_DAY_STAT[i]
        sc_letter = col_letter(sc)
        R.append(conditional_format_request(
            sheet_id, 6, 13, LIVE_DAY_STAT[i], LIVE_DAY_TASK[i] + 1,
            formula=f"=${sc_letter}7=TRUE",
            bg_hex=SAGE_EXTRA_LIGHT_HEX,
            text_hex=MUTED_HEX,
            strikethrough=True,
        ))

    # Row 15 (0-based 14): frozen day-of-week headers, merged 2-wide each
    for i, d in enumerate(WEEKDAY_NAMES):
        R.append(merge_cells_request(sheet_id, 14, 15, LIVE_DAY_STAT[i], LIVE_DAY_TASK[i] + 1))
        V.append(set_value_request(sheet_id, 14, LIVE_DAY_STAT[i], d.upper(), fmt={
            "horizontalAlignment": "CENTER",
            "verticalAlignment": "MIDDLE",
            "backgroundColor": hex_to_rgb(SAGE_DARK_HEX),
            "textFormat": {"bold": True, "fontSize": 11, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
        }, formula=False))

    # Rows 17-21 (0-based 16..20): Big % display per day, merged 2-wide
    # Anchored at row 16 (0-based) — actually use row 16 only and leave 17..20 as visual spacer per spec.
    # Spec said rows 17-21. We'll merge rows 17-21 (0-based 16..21) across each day-pair into ONE big cell.
    for i in range(7):
        R.append(merge_cells_request(sheet_id, 16, 21, LIVE_DAY_STAT[i], LIVE_DAY_TASK[i] + 1))
        sc_letter = col_letter(LIVE_DAY_STAT[i])
        tc_letter = col_letter(LIVE_DAY_TASK[i])
        # The slot status range is sc23:sc37 (0-based 22..36) ; task range tc23:tc37
        # COUNTIF on TRUE checkboxes, COUNTA on task text. If task range empty → show em-dash.
        formula = (
            f'=IF(COUNTA({tc_letter}23:{tc_letter}37)=0,"—",'
            f'COUNTIF({sc_letter}23:{sc_letter}37,TRUE)/COUNTA({tc_letter}23:{tc_letter}37))'
        )
        V.append(set_value_request(sheet_id, 16, LIVE_DAY_STAT[i], formula, fmt={
            "horizontalAlignment": "CENTER",
            "verticalAlignment": "MIDDLE",
            "backgroundColor": hex_to_rgb(SAGE_LIGHT_HEX),
            "numberFormat": {"type": "PERCENT", "pattern": "0%"},
            "textFormat": {"bold": True, "fontSize": 22, "foregroundColor": hex_to_rgb(SAGE_DARK_HEX)},
        }, formula=True))

    # Rows 23-37 (0-based 22..36): 15 priority-slot rows. Status checkbox + task text per day.
    for i in range(7):
        sc = LIVE_DAY_STAT[i]
        tc = LIVE_DAY_TASK[i]
        R.append(checkbox_validation_request(sheet_id, 22, 37, sc, sc + 1))
        # Task col format: align left, small font
        R.append(format_range_request(sheet_id, 22, 37, tc, tc + 1, {
            "horizontalAlignment": "LEFT",
            "verticalAlignment": "MIDDLE",
            "textFormat": {"fontSize": 10, "foregroundColor": hex_to_rgb(INK_HEX)},
            "wrapStrategy": "WRAP",
        }, fields="userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat,wrapStrategy)"))
        # Conditional formatting: status TRUE → strikethrough + sage-extra-light fill
        sc_letter = col_letter(sc)
        R.append(conditional_format_request(
            sheet_id, 22, 37, sc, tc + 1,
            formula=f"=${sc_letter}23=TRUE",
            bg_hex=SAGE_EXTRA_LIGHT_HEX,
            text_hex=MUTED_HEX,
            strikethrough=True,
        ))

    # Row 39 (0-based 38): "notes · ideas · journal" label
    for i in range(7):
        R.append(merge_cells_request(sheet_id, 38, 39, LIVE_DAY_STAT[i], LIVE_DAY_TASK[i] + 1))
        V.append(set_value_request(sheet_id, 38, LIVE_DAY_STAT[i], "notes · ideas · journal", fmt={
            "horizontalAlignment": "CENTER",
            "textFormat": {"italic": True, "fontSize": 9, "foregroundColor": hex_to_rgb(MUTED_HEX)},
        }, formula=False))

    # Rows 40-47 (0-based 39..46): notes block — wrap text, left align
    for i in range(7):
        R.append(merge_cells_request(sheet_id, 39, 47, LIVE_DAY_STAT[i], LIVE_DAY_TASK[i] + 1))
        R.append(format_range_request(sheet_id, 39, 47, LIVE_DAY_STAT[i], LIVE_DAY_TASK[i] + 1, {
            "horizontalAlignment": "LEFT",
            "verticalAlignment": "TOP",
            "wrapStrategy": "WRAP",
            "textFormat": {"fontSize": 9, "foregroundColor": hex_to_rgb(INK_HEX)},
        }, fields="userEnteredFormat(horizontalAlignment,verticalAlignment,wrapStrategy,textFormat)"))

    return R, V


# ---------------------------------------------------------------- To Do tab

# To Do columns (0-based)
TODO_COL_STATUS = 0  # A — checkbox
TODO_COL_TASK = 1    # B
TODO_COL_TYPE = 2    # C
TODO_COL_PROJECT = 3  # D
TODO_COL_DUE = 4     # E
TODO_COL_NOTES = 5   # F

TODO_HEADERS = ["Status", "Task", "Type", "Project", "Due", "Notes"]

PROJECT_OPTIONS = ["G&B", "Kai Grey", "Panthera Grey", "Myself Renewed", "Home"]
TYPE_OPTIONS = ["Work", "Home"]

TODO_MAX_ROWS = 200


def build_todo_requests(sheet_id: int) -> tuple[list[dict], list[dict]]:
    R: list[dict] = []
    V: list[dict] = []

    # Column widths
    widths = {
        TODO_COL_STATUS: 50,
        TODO_COL_TASK: 360,
        TODO_COL_TYPE: 90,
        TODO_COL_PROJECT: 130,
        TODO_COL_DUE: 100,
        TODO_COL_NOTES: 280,
    }
    for col, w in widths.items():
        R.append(update_dim_request(sheet_id, "COLUMNS", col, col + 1, w))
    R.append(update_dim_request(sheet_id, "ROWS", 0, 1, 32))

    # Header row 1 (0-based row 0)
    for c, header in enumerate(TODO_HEADERS):
        V.append(set_value_request(sheet_id, 0, c, header, fmt={
            "horizontalAlignment": "LEFT" if c != TODO_COL_STATUS else "CENTER",
            "verticalAlignment": "MIDDLE",
            "backgroundColor": hex_to_rgb(SAGE_DARK_HEX),
            "textFormat": {"bold": True, "fontSize": 10, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
        }, formula=False))

    # Status checkboxes across data rows
    R.append(checkbox_validation_request(sheet_id, 1, TODO_MAX_ROWS, TODO_COL_STATUS, TODO_COL_STATUS + 1))
    # Type dropdown
    R.append(dropdown_validation_request(sheet_id, 1, TODO_MAX_ROWS, TODO_COL_TYPE, TODO_COL_TYPE + 1, TYPE_OPTIONS, strict=True))
    # Project dropdown (non-strict so free text is OK)
    R.append(dropdown_validation_request(sheet_id, 1, TODO_MAX_ROWS, TODO_COL_PROJECT, TODO_COL_PROJECT + 1, PROJECT_OPTIONS, strict=False))

    # Task col + notes col wrap
    R.append(format_range_request(sheet_id, 1, TODO_MAX_ROWS, TODO_COL_TASK, TODO_COL_TASK + 1, {
        "wrapStrategy": "WRAP", "verticalAlignment": "MIDDLE",
        "textFormat": {"fontSize": 10, "foregroundColor": hex_to_rgb(INK_HEX)},
    }, fields="userEnteredFormat(wrapStrategy,verticalAlignment,textFormat)"))
    R.append(format_range_request(sheet_id, 1, TODO_MAX_ROWS, TODO_COL_NOTES, TODO_COL_NOTES + 1, {
        "wrapStrategy": "WRAP", "verticalAlignment": "MIDDLE",
        "textFormat": {"fontSize": 9, "foregroundColor": hex_to_rgb(MUTED_HEX)},
    }, fields="userEnteredFormat(wrapStrategy,verticalAlignment,textFormat)"))

    # Due date column format
    R.append(format_range_request(sheet_id, 1, TODO_MAX_ROWS, TODO_COL_DUE, TODO_COL_DUE + 1, {
        "numberFormat": {"type": "DATE", "pattern": "yyyy-mm-dd"},
        "horizontalAlignment": "CENTER",
        "textFormat": {"fontSize": 10},
    }, fields="userEnteredFormat(numberFormat,horizontalAlignment,textFormat)"))

    # Conditional formatting: Status TRUE → strikethrough + sage-light fill across row
    R.append(conditional_format_request(
        sheet_id, 1, TODO_MAX_ROWS, 0, len(TODO_HEADERS),
        formula="=$A2=TRUE",
        bg_hex=SAGE_EXTRA_LIGHT_HEX,
        text_hex=MUTED_HEX,
        strikethrough=True,
    ))
    # Type column tinting
    R.append(conditional_format_request(
        sheet_id, 1, TODO_MAX_ROWS, TODO_COL_TYPE, TODO_COL_TYPE + 1,
        formula='=$C2="Home"', bg_hex=TYPE_HOME_HEX, index=0,
    ))
    R.append(conditional_format_request(
        sheet_id, 1, TODO_MAX_ROWS, TODO_COL_TYPE, TODO_COL_TYPE + 1,
        formula='=$C2="Work"', bg_hex=TYPE_WORK_HEX, index=0,
    ))

    return R, V


# ---------------------------------------------------------------- To Do Long Term tab

# Same column structure as To Do, but Status is a dropdown (not checkbox).
LT_COL_STATUS = 0  # A — dropdown (Idea/Active/On hold/Promoted/Done)
LT_COL_TASK = 1
LT_COL_TYPE = 2
LT_COL_PROJECT = 3
LT_COL_DUE = 4
LT_COL_NOTES = 5

LT_HEADERS = ["Status", "Item", "Type", "Project", "Due", "Notes"]
LT_STATUS_OPTIONS = ["Idea", "Active", "On hold", "Promoted", "Done"]
LT_MAX_ROWS = 200


def build_lt_requests(sheet_id: int) -> tuple[list[dict], list[dict]]:
    R: list[dict] = []
    V: list[dict] = []

    widths = {
        LT_COL_STATUS: 100,
        LT_COL_TASK: 360,
        LT_COL_TYPE: 90,
        LT_COL_PROJECT: 130,
        LT_COL_DUE: 100,
        LT_COL_NOTES: 280,
    }
    for col, w in widths.items():
        R.append(update_dim_request(sheet_id, "COLUMNS", col, col + 1, w))
    R.append(update_dim_request(sheet_id, "ROWS", 0, 1, 32))

    for c, header in enumerate(LT_HEADERS):
        V.append(set_value_request(sheet_id, 0, c, header, fmt={
            "horizontalAlignment": "LEFT",
            "verticalAlignment": "MIDDLE",
            "backgroundColor": hex_to_rgb(SAGE_DARK_HEX),
            "textFormat": {"bold": True, "fontSize": 10, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
        }, formula=False))

    R.append(dropdown_validation_request(sheet_id, 1, LT_MAX_ROWS, LT_COL_STATUS, LT_COL_STATUS + 1, LT_STATUS_OPTIONS, strict=True))
    R.append(dropdown_validation_request(sheet_id, 1, LT_MAX_ROWS, LT_COL_TYPE, LT_COL_TYPE + 1, TYPE_OPTIONS, strict=True))
    R.append(dropdown_validation_request(sheet_id, 1, LT_MAX_ROWS, LT_COL_PROJECT, LT_COL_PROJECT + 1, PROJECT_OPTIONS, strict=False))

    R.append(format_range_request(sheet_id, 1, LT_MAX_ROWS, LT_COL_TASK, LT_COL_TASK + 1, {
        "wrapStrategy": "WRAP", "verticalAlignment": "MIDDLE",
        "textFormat": {"fontSize": 10, "foregroundColor": hex_to_rgb(INK_HEX)},
    }, fields="userEnteredFormat(wrapStrategy,verticalAlignment,textFormat)"))

    R.append(format_range_request(sheet_id, 1, LT_MAX_ROWS, LT_COL_DUE, LT_COL_DUE + 1, {
        "numberFormat": {"type": "DATE", "pattern": "yyyy-mm-dd"},
        "horizontalAlignment": "CENTER",
    }, fields="userEnteredFormat(numberFormat,horizontalAlignment)"))

    # Done → strikethrough + sage-extra-light
    R.append(conditional_format_request(
        sheet_id, 1, LT_MAX_ROWS, 0, len(LT_HEADERS),
        formula='=$A2="Done"',
        bg_hex=SAGE_EXTRA_LIGHT_HEX,
        text_hex=MUTED_HEX,
        strikethrough=True,
    ))
    # Promoted → muted grey
    R.append(conditional_format_request(
        sheet_id, 1, LT_MAX_ROWS, 0, len(LT_HEADERS),
        formula='=$A2="Promoted"',
        bg_hex="ececec",
        text_hex=MUTED_HEX,
        italic=True,
    ))

    return R, V


# ---------------------------------------------------------------- Projects tab

PJ_HEADERS = ["Project", "Entity", "Status", "Start", "Target", "Tab", "Notes"]
PJ_COL_PROJECT, PJ_COL_ENTITY, PJ_COL_STATUS, PJ_COL_START, PJ_COL_TARGET, PJ_COL_TAB, PJ_COL_NOTES = range(7)
PJ_STATUS_OPTIONS = ["Plan Needed", "Active", "On hold", "Done"]
PJ_MAX_ROWS = 50


def build_projects_requests(sheet_id: int) -> tuple[list[dict], list[dict]]:
    R: list[dict] = []
    V: list[dict] = []

    widths = {
        PJ_COL_PROJECT: 260,
        PJ_COL_ENTITY: 130,
        PJ_COL_STATUS: 110,
        PJ_COL_START: 100,
        PJ_COL_TARGET: 100,
        PJ_COL_TAB: 80,
        PJ_COL_NOTES: 380,
    }
    for col, w in widths.items():
        R.append(update_dim_request(sheet_id, "COLUMNS", col, col + 1, w))
    R.append(update_dim_request(sheet_id, "ROWS", 0, 1, 32))

    for c, header in enumerate(PJ_HEADERS):
        V.append(set_value_request(sheet_id, 0, c, header, fmt={
            "horizontalAlignment": "LEFT",
            "verticalAlignment": "MIDDLE",
            "backgroundColor": hex_to_rgb(SAGE_DARK_HEX),
            "textFormat": {"bold": True, "fontSize": 10, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
        }, formula=False))

    R.append(dropdown_validation_request(sheet_id, 1, PJ_MAX_ROWS, PJ_COL_ENTITY, PJ_COL_ENTITY + 1, PROJECT_OPTIONS, strict=True))
    R.append(dropdown_validation_request(sheet_id, 1, PJ_MAX_ROWS, PJ_COL_STATUS, PJ_COL_STATUS + 1, PJ_STATUS_OPTIONS, strict=True))

    R.append(format_range_request(sheet_id, 1, PJ_MAX_ROWS, PJ_COL_START, PJ_COL_TARGET + 1, {
        "numberFormat": {"type": "DATE", "pattern": "yyyy-mm-dd"},
        "horizontalAlignment": "CENTER",
    }, fields="userEnteredFormat(numberFormat,horizontalAlignment)"))

    # Tab column: hyperlinks
    R.append(format_range_request(sheet_id, 1, PJ_MAX_ROWS, PJ_COL_TAB, PJ_COL_TAB + 1, {
        "horizontalAlignment": "CENTER",
        "textFormat": {"foregroundColor": hex_to_rgb(SAGE_DARK_HEX), "underline": True, "fontSize": 10},
    }, fields="userEnteredFormat(horizontalAlignment,textFormat)"))

    # Done → strikethrough
    R.append(conditional_format_request(
        sheet_id, 1, PJ_MAX_ROWS, 0, len(PJ_HEADERS),
        formula='=$C2="Done"',
        bg_hex=SAGE_EXTRA_LIGHT_HEX,
        text_hex=MUTED_HEX,
        strikethrough=True,
    ))
    # Entity column tinting
    for ent, hexc in ENTITY_COLOR_HEX.items():
        R.append(conditional_format_request(
            sheet_id, 1, PJ_MAX_ROWS, PJ_COL_ENTITY, PJ_COL_ENTITY + 1,
            formula=f'=$B2="{ent}"',
            bg_hex=hexc,
            index=0,
        ))

    return R, V


# ---------------------------------------------------------------- Gantt tab builder

GANTT_FIRST_WEEK_COL = 5  # 0-based: after Status(0), Milestone(1), Start(2), Target(3), Notes(4)
GANTT_HEADERS_FIXED = ["Status", "Milestone", "Start", "Target", "Notes"]


def build_gantt_requests(sheet_id: int, project_name: str, entity: str,
                         start_iso: str, weeks: int = 16) -> tuple[list[dict], list[dict]]:
    R: list[dict] = []
    V: list[dict] = []

    # Title row 2 (0-based 1)
    last_col = GANTT_FIRST_WEEK_COL + weeks  # exclusive
    R.append(merge_cells_request(sheet_id, 1, 2, 0, last_col))
    V.append(set_value_request(sheet_id, 1, 0, project_name.upper(), fmt={
        "horizontalAlignment": "LEFT", "verticalAlignment": "MIDDLE",
        "backgroundColor": hex_to_rgb(SAGE_LIGHT_HEX),
        "textFormat": {"bold": True, "fontSize": 16, "foregroundColor": hex_to_rgb(SAGE_DARK_HEX)},
    }, formula=False))
    R.append(update_dim_request(sheet_id, "ROWS", 1, 2, 36))

    # Subtitle row 3 (0-based 2)
    R.append(merge_cells_request(sheet_id, 2, 3, 0, last_col))
    V.append(set_value_request(sheet_id, 2, 0,
        f"Entity: {entity}  ·  Tick the week boxes you're actively working on each milestone — the row builds into a Gantt bar",
        fmt={
            "horizontalAlignment": "LEFT", "verticalAlignment": "MIDDLE",
            "textFormat": {"italic": True, "fontSize": 9, "foregroundColor": hex_to_rgb(MUTED_HEX)},
        }, formula=False))

    # Header row 5 (0-based 4)
    for c, header in enumerate(GANTT_HEADERS_FIXED):
        V.append(set_value_request(sheet_id, 4, c, header, fmt={
            "horizontalAlignment": "LEFT" if c > 0 else "CENTER",
            "verticalAlignment": "MIDDLE",
            "backgroundColor": hex_to_rgb(SAGE_DARK_HEX),
            "textFormat": {"bold": True, "fontSize": 10, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
        }, formula=False))

    # Week column headers
    y, m, d = (int(x) for x in start_iso.split("-"))
    start_dt = date(y, m, d)
    monday = start_dt - timedelta(days=start_dt.weekday())
    for w in range(weeks):
        wk = monday + timedelta(days=7 * w)
        col = GANTT_FIRST_WEEK_COL + w
        V.append(set_value_request(sheet_id, 4, col, f"{wk.month}/{wk.day}", fmt={
            "horizontalAlignment": "CENTER",
            "verticalAlignment": "MIDDLE",
            "backgroundColor": hex_to_rgb(SAGE_DARK_HEX),
            "textFormat": {"bold": True, "fontSize": 9, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
        }, formula=False))

    # Column widths
    R.append(update_dim_request(sheet_id, "COLUMNS", 0, 1, 50))   # status
    R.append(update_dim_request(sheet_id, "COLUMNS", 1, 2, 360))  # milestone
    R.append(update_dim_request(sheet_id, "COLUMNS", 2, 3, 100))  # start
    R.append(update_dim_request(sheet_id, "COLUMNS", 3, 4, 100))  # target
    R.append(update_dim_request(sheet_id, "COLUMNS", 4, 5, 260))  # notes
    for w in range(weeks):
        col = GANTT_FIRST_WEEK_COL + w
        R.append(update_dim_request(sheet_id, "COLUMNS", col, col + 1, 44))

    # Status + week checkboxes — milestone rows are 6..15 (0-based 5..14)
    R.append(checkbox_validation_request(sheet_id, 5, 15, 0, 1))
    R.append(checkbox_validation_request(sheet_id, 5, 15, GANTT_FIRST_WEEK_COL, last_col))

    # Date format on Start/Target
    R.append(format_range_request(sheet_id, 5, 15, 2, 4, {
        "numberFormat": {"type": "DATE", "pattern": "yyyy-mm-dd"},
        "horizontalAlignment": "CENTER",
    }, fields="userEnteredFormat(numberFormat,horizontalAlignment)"))
    R.append(format_range_request(sheet_id, 5, 15, 1, 2, {
        "wrapStrategy": "WRAP", "verticalAlignment": "MIDDLE",
        "textFormat": {"fontSize": 10, "foregroundColor": hex_to_rgb(INK_HEX)},
    }, fields="userEnteredFormat(wrapStrategy,verticalAlignment,textFormat)"))
    R.append(format_range_request(sheet_id, 5, 15, 4, 5, {
        "wrapStrategy": "WRAP", "verticalAlignment": "MIDDLE",
        "textFormat": {"fontSize": 9, "foregroundColor": hex_to_rgb(MUTED_HEX)},
    }, fields="userEnteredFormat(wrapStrategy,verticalAlignment,textFormat)"))

    # Conditional formatting: week-cell TRUE → entity-color fill
    entity_hex = ENTITY_COLOR_HEX.get(entity, SAGE_DARK_HEX)
    first_week_letter = col_letter(GANTT_FIRST_WEEK_COL)
    R.append(conditional_format_request(
        sheet_id, 5, 15, GANTT_FIRST_WEEK_COL, last_col,
        formula=f"={first_week_letter}6=TRUE",
        bg_hex=entity_hex,
    ))
    # Status TRUE → milestone row strikethrough + muted grey
    R.append(conditional_format_request(
        sheet_id, 5, 15, 0, 5,
        formula="=$A6=TRUE",
        bg_hex=SAGE_EXTRA_LIGHT_HEX,
        text_hex=MUTED_HEX,
        strikethrough=True,
    ))

    return R, V


# ---------------------------------------------------------------- main builder

def main() -> int:
    today = date.today()
    week_label = live_week_label(today)
    live_tab_title = week_label

    print(f"[1/8] Auth …")
    token = get_access_token()
    sheets = SheetsClient(token)

    print(f"[2/8] Create spreadsheet '{NEW_SHEET_TITLE}' …")
    ss = sheets.create_spreadsheet(NEW_SHEET_TITLE)
    sid = ss["spreadsheetId"]
    sheet_url = ss["spreadsheetUrl"]
    default_sheet_id = ss["sheets"][0]["properties"]["sheetId"]
    print(f"      → id={sid}")
    print(f"      → url={sheet_url}")

    print(f"[3/8] Move to STRATEGIC PLANNING folder …")
    move_to_folder(token, sid, PARENT_FOLDER_ID)

    print(f"[4/8] Create 5 tabs + rename default …")
    # First, rename the default tab to the live-week label, then create 4 more tabs.
    init_reqs = [
        {"updateSheetProperties": {
            "properties": {
                "sheetId": default_sheet_id,
                "title": live_tab_title,
                "gridProperties": {
                    "rowCount": 80,
                    "columnCount": 15,
                    "frozenRowCount": 15,
                },
                "tabColor": hex_to_rgb(SAGE_DARK_HEX),
            },
            "fields": "title,gridProperties.rowCount,gridProperties.columnCount,gridProperties.frozenRowCount,tabColor",
        }},
        # We'll create the other tabs in the next batch and capture their sheetIds via the response
    ]
    resp = sheets.batch_update(sid, init_reqs)

    # Now add the other 4 tabs.
    add_tabs = [
        {"addSheet": {"properties": {
            "title": "To Do",
            "gridProperties": {"rowCount": TODO_MAX_ROWS, "columnCount": 8, "frozenRowCount": 1},
        }}},
        {"addSheet": {"properties": {
            "title": "To Do Long Term",
            "gridProperties": {"rowCount": LT_MAX_ROWS, "columnCount": 8, "frozenRowCount": 1},
        }}},
        {"addSheet": {"properties": {
            "title": "Projects",
            "gridProperties": {"rowCount": PJ_MAX_ROWS, "columnCount": 10, "frozenRowCount": 1},
        }}},
        {"addSheet": {"properties": {
            "title": "Myself Renewed Healthcare",
            "gridProperties": {"rowCount": 30, "columnCount": 24, "frozenRowCount": 5},
        }}},
    ]
    resp = sheets.batch_update(sid, add_tabs)
    new_sheet_ids: dict[str, int] = {}
    for r in resp.get("replies", []):
        p = r.get("addSheet", {}).get("properties", {})
        new_sheet_ids[p["title"]] = p["sheetId"]
    live_sid = default_sheet_id
    todo_sid = new_sheet_ids["To Do"]
    lt_sid = new_sheet_ids["To Do Long Term"]
    pj_sid = new_sheet_ids["Projects"]
    healthcare_sid = new_sheet_ids["Myself Renewed Healthcare"]

    print(f"      → live={live_sid}, todo={todo_sid}, lt={lt_sid}, projects={pj_sid}, healthcare={healthcare_sid}")

    print(f"[5/8] Build Live Week structure …")
    live_R, live_V = build_live_week_requests(live_sid, week_label)
    sheets.batch_update(sid, live_R)
    sheets.batch_update(sid, live_V)

    print(f"[6/8] Build To Do, To Do Long Term, Projects, Healthcare …")
    for label, build_fn, ssid in [
        ("To Do", build_todo_requests, todo_sid),
        ("To Do Long Term", build_lt_requests, lt_sid),
        ("Projects", build_projects_requests, pj_sid),
    ]:
        R, V = build_fn(ssid)
        # Split large batches if needed
        sheets.batch_update(sid, R)
        sheets.batch_update(sid, V)

    # Healthcare Gantt — 16 weeks from 2026-04-27 (read from source row 6 Start)
    gantt_R, gantt_V = build_gantt_requests(healthcare_sid, "Myself Renewed Healthcare",
                                             "Myself Renewed", "2026-04-27", weeks=16)
    sheets.batch_update(sid, gantt_R)
    sheets.batch_update(sid, gantt_V)

    # Also add Deal Aggregator Expansion Gantt tab (12 weeks from 2026-05-11)
    add_da = [{"addSheet": {"properties": {
        "title": "Deal Aggregator Expansion",
        "gridProperties": {"rowCount": 20, "columnCount": 20, "frozenRowCount": 5},
    }}}]
    resp = sheets.batch_update(sid, add_da)
    da_sid = resp["replies"][0]["addSheet"]["properties"]["sheetId"]
    da_R, da_V = build_gantt_requests(da_sid, "Deal Aggregator Expansion", "G&B", "2026-05-11", weeks=12)
    sheets.batch_update(sid, da_R)
    sheets.batch_update(sid, da_V)

    print(f"[7/8] Migrate data from source Excel …")
    migrate_data(token, sid, live_sid, todo_sid, lt_sid, pj_sid, healthcare_sid, da_sid, today)

    print(f"[8/8] Done.")
    print()
    print(f"NEW_SHEET_ID={sid}")
    print(f"NEW_SHEET_URL={sheet_url}")
    return 0


# ---------------------------------------------------------------- Phase 2 data migration

def migrate_data(token: str, sid: str, live_sid: int, todo_sid: int, lt_sid: int,
                 pj_sid: int, healthcare_sid: int, da_sid: int, today: date) -> None:
    wb = load_workbook(SOURCE_XLSX_PATH, data_only=False)
    sheets = SheetsClient(token)

    # ---------- Live Week: habits + priorities + notes
    src_live = wb["May 11-17"]
    live_R: list[dict] = []

    # Habits — source rows 7..13 cols B/D/F/H/J/L/N (excel B=2)
    src_day_cols = [2, 4, 6, 8, 10, 12, 14]  # B,D,F,H,J,L,N (1-based)
    for h_idx in range(7):
        src_row = 7 + h_idx
        dst_row = 6 + h_idx  # 0-based
        for i in range(7):
            val = src_live.cell(src_row, src_day_cols[i]).value
            checked = val == "✅"
            live_R.append(set_value_request(live_sid, dst_row, LIVE_DAY_STAT[i], checked))

    # Priorities — source rows 23..37: status in B/D/F/H/J/L/N, task in C/E/G/I/K/M/O
    for slot in range(15):
        src_row = 23 + slot
        dst_row = 22 + slot
        for i in range(7):
            stat_col = src_day_cols[i]  # 1-based
            task_col = stat_col + 1
            stat_val = src_live.cell(src_row, stat_col).value
            task_val = src_live.cell(src_row, task_col).value
            checked = stat_val == "✅"
            live_R.append(set_value_request(live_sid, dst_row, LIVE_DAY_STAT[i], checked))
            if task_val:
                live_R.append(set_value_request(live_sid, dst_row, LIVE_DAY_TASK[i], task_val))

    # Notes rows 40..47 → 0-based 39..46 (merged across day-pair, anchor is stat col)
    for n_idx in range(8):
        src_row = 40 + n_idx
        dst_row = 39 + n_idx
        for i in range(7):
            stat_col = src_day_cols[i]
            note_val = src_live.cell(src_row, stat_col).value
            if note_val:
                live_R.append(set_value_request(live_sid, dst_row, LIVE_DAY_STAT[i], note_val))

    # Batch live writes
    _batch_in_chunks(sheets, sid, live_R, chunk=200)

    # ---------- To Do
    src_td = wb["To Do"]
    todo_R: list[dict] = []
    dst_row = 1  # 0-based row after header
    for r in range(6, src_td.max_row + 1):
        task = src_td.cell(r, 3).value  # col C
        if not task:
            continue
        status = src_td.cell(r, 2).value  # B
        type_ = src_td.cell(r, 4).value or "Work"  # D
        project = src_td.cell(r, 5).value or ""  # E
        due = src_td.cell(r, 6).value or ""  # F
        notes = src_td.cell(r, 7).value or ""  # G
        checked = status == "✅"
        # Normalize "n/a" project → blank
        if project == "n/a":
            project = ""
        # Normalize Myself Renewed Healthcare → Myself Renewed for dropdown match
        if project == "Myself Renewed Healthcare":
            project = "Myself Renewed"
        # Personal → Home
        if project == "Personal":
            project = "Home"
        # Type fallback
        if type_ not in ("Work", "Home"):
            type_ = "Work"

        todo_R.append(set_value_request(todo_sid, dst_row, TODO_COL_STATUS, checked))
        todo_R.append(set_value_request(todo_sid, dst_row, TODO_COL_TASK, str(task)))
        todo_R.append(set_value_request(todo_sid, dst_row, TODO_COL_TYPE, type_))
        if project:
            todo_R.append(set_value_request(todo_sid, dst_row, TODO_COL_PROJECT, project))
        if due:
            todo_R.append(set_value_request(todo_sid, dst_row, TODO_COL_DUE, str(due)[:10]))
        if notes:
            todo_R.append(set_value_request(todo_sid, dst_row, TODO_COL_NOTES, str(notes)))
        dst_row += 1

    # Append today's Saltoun row
    saltoun_row = dst_row
    todo_R.append(set_value_request(todo_sid, dst_row, TODO_COL_STATUS, False))
    todo_R.append(set_value_request(todo_sid, dst_row, TODO_COL_TASK,
        "Reply to Janet Crockett — confirm year-end 12/31/2025 G&B investment figures (Saltoun annual review)"))
    todo_R.append(set_value_request(todo_sid, dst_row, TODO_COL_TYPE, "Work"))
    todo_R.append(set_value_request(todo_sid, dst_row, TODO_COL_PROJECT, "G&B"))
    todo_R.append(set_value_request(todo_sid, dst_row, TODO_COL_DUE, today.isoformat()))
    todo_R.append(set_value_request(todo_sid, dst_row, TODO_COL_NOTES,
        "Citrin Cooperman accountants doing annual portfolio review. Inbox: brain/inbox/2026-05-12-saltoun-annual-financial-review.md"))
    dst_row += 1

    _batch_in_chunks(sheets, sid, todo_R, chunk=200)
    print(f"      → migrated To Do rows + appended Saltoun at row {saltoun_row + 1} (1-based)")

    # ---------- To Do Long Term
    src_lt = wb["To Do Long Term"]
    lt_R: list[dict] = []
    dst_row = 1
    for r in range(6, src_lt.max_row + 1):
        item = src_lt.cell(r, 3).value
        if not item:
            continue
        status = src_lt.cell(r, 5).value or "Idea"  # E
        type_ = src_lt.cell(r, 4).value or "Home"
        if status not in LT_STATUS_OPTIONS:
            status = "Idea"
        if type_ not in TYPE_OPTIONS:
            type_ = "Home"
        lt_R.append(set_value_request(lt_sid, dst_row, LT_COL_STATUS, status))
        lt_R.append(set_value_request(lt_sid, dst_row, LT_COL_TASK, str(item)))
        lt_R.append(set_value_request(lt_sid, dst_row, LT_COL_TYPE, type_))
        notes = src_lt.cell(r, 6).value
        if notes:
            lt_R.append(set_value_request(lt_sid, dst_row, LT_COL_NOTES, str(notes)))
        dst_row += 1
    _batch_in_chunks(sheets, sid, lt_R, chunk=200)

    # ---------- Projects index (with hyperlinks to tabs)
    src_pj = wb["Projects"]
    pj_R: list[dict] = []
    dst_row = 1
    # We need the gid of each Gantt tab to build hyperlinks.
    gantt_gids = {"Myself Renewed Healthcare": healthcare_sid, "Deal Aggregator Expansion": da_sid}
    for r in range(6, src_pj.max_row + 1):
        name = src_pj.cell(r, 2).value  # B
        if not name:
            continue
        entity = src_pj.cell(r, 3).value or "G&B"
        status = src_pj.cell(r, 4).value or "Active"
        start = src_pj.cell(r, 5).value or ""
        target = src_pj.cell(r, 6).value or ""
        notes = src_pj.cell(r, 8).value or ""
        if entity == "Myself Renewed Healthcare":
            entity = "Myself Renewed"
        if status not in PJ_STATUS_OPTIONS:
            status = "Active"
        pj_R.append(set_value_request(pj_sid, dst_row, PJ_COL_PROJECT, str(name)))
        pj_R.append(set_value_request(pj_sid, dst_row, PJ_COL_ENTITY, str(entity)))
        pj_R.append(set_value_request(pj_sid, dst_row, PJ_COL_STATUS, str(status)))
        if start:
            pj_R.append(set_value_request(pj_sid, dst_row, PJ_COL_START, str(start)[:10]))
        if target:
            pj_R.append(set_value_request(pj_sid, dst_row, PJ_COL_TARGET, str(target)[:10]))
        gid = gantt_gids.get(name)
        if gid is not None:
            link = f'=HYPERLINK("#gid={gid}","Open")'
            pj_R.append(set_value_request(pj_sid, dst_row, PJ_COL_TAB, link, formula=True))
        if notes:
            pj_R.append(set_value_request(pj_sid, dst_row, PJ_COL_NOTES, str(notes)))
        dst_row += 1
    _batch_in_chunks(sheets, sid, pj_R, chunk=200)

    # ---------- Myself Renewed Healthcare (milestones + ticks)
    src_hc = wb["Myself Renewed Healthcare"]
    hc_R: list[dict] = []
    # Source: rows 6..15, cols C(milestone), D(start), E(target), F(notes), G..V (16 week cells)
    for m_idx in range(10):
        src_row = 6 + m_idx
        dst_row = 5 + m_idx  # 0-based
        milestone = src_hc.cell(src_row, 3).value
        if not milestone:
            continue
        start = src_hc.cell(src_row, 4).value
        target = src_hc.cell(src_row, 5).value
        notes = src_hc.cell(src_row, 6).value
        status_val = src_hc.cell(src_row, 2).value
        hc_R.append(set_value_request(healthcare_sid, dst_row, 0, status_val == "✅"))
        hc_R.append(set_value_request(healthcare_sid, dst_row, 1, str(milestone)))
        if start:
            hc_R.append(set_value_request(healthcare_sid, dst_row, 2, str(start)[:10]))
        if target:
            hc_R.append(set_value_request(healthcare_sid, dst_row, 3, str(target)[:10]))
        if notes:
            hc_R.append(set_value_request(healthcare_sid, dst_row, 4, str(notes)))
        # Week cells: source G..V (cols 7..22, 1-based), dst cols 5..20 (0-based)
        for w in range(16):
            v = src_hc.cell(src_row, 7 + w).value
            checked = v == "✅"
            hc_R.append(set_value_request(healthcare_sid, dst_row, GANTT_FIRST_WEEK_COL + w, checked))
    _batch_in_chunks(sheets, sid, hc_R, chunk=200)

    # ---------- Deal Aggregator Expansion (mostly empty — copy ticks if any)
    src_da = wb["Deal Aggregator Expansion"]
    da_R: list[dict] = []
    for m_idx in range(10):
        src_row = 6 + m_idx
        dst_row = 5 + m_idx
        milestone = src_da.cell(src_row, 3).value
        if not milestone:
            continue
        da_R.append(set_value_request(da_sid, dst_row, 1, str(milestone)))
        status_val = src_da.cell(src_row, 2).value
        da_R.append(set_value_request(da_sid, dst_row, 0, status_val == "✅"))
    _batch_in_chunks(sheets, sid, da_R, chunk=200)

    # ---------- Promote Saltoun to today's slot 1
    # Today is 2026-05-12 = Tuesday (weekday=1)
    today_dow = today.weekday()  # 0=Mon
    saltoun_task = ("Reply to Janet Crockett — confirm year-end 12/31/2025 G&B investment "
                    "figures (Saltoun annual review)")
    promote_R = [
        set_value_request(live_sid, 22, LIVE_DAY_TASK[today_dow], saltoun_task),  # slot 1 = row 23 (0-based 22)
        set_value_request(live_sid, 22, LIVE_DAY_STAT[today_dow], False),
    ]
    sheets.batch_update(sid, promote_R)
    print(f"      → promoted Saltoun to today ({['Mon','Tue','Wed','Thu','Fri','Sat','Sun'][today_dow]}) slot 1")


def _batch_in_chunks(sheets: SheetsClient, sid: str, reqs: list[dict], chunk: int = 200) -> None:
    for i in range(0, len(reqs), chunk):
        sheets.batch_update(sid, reqs[i:i + chunk])


if __name__ == "__main__":
    sys.exit(main())
