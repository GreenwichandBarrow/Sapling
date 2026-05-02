"""One-shot maintenance pass on Kay's live TO DO xlsx.

Does NOT rebuild the file (would wipe data). Only:
  1. Removes donut chart objects from the This Week tab.
  2. Promotes the per-day % display into the now-empty donut area
     (rows 17-21 merged cell) at large size; clears the redundant
     row 22 percentage row.
  3. Re-applies the strike+green conditional formatting on:
     - This Week priorities (B23:O37 in 7 day-pair ranges)
     - This Week habits (B7:N13)
     - To Do tab (B6:G85)
     - To Do Long Term tab (B6:F40)
     - Myself Renewed Healthcare Gantt (milestone left-side + timeline)
  4. Renames the "This Week" tab to the current Mon-Sun week range
     (e.g. "Apr 27-May 3") so Kay can see the week at the bottom of
     Excel without opening the tab. Date formulas inside the tab
     keep working because they reference TODAY(), not the tab name.

Run AFTER closing Excel:
    python3 scripts/maintain_tasks_excel.py

Idempotent — safe to run more than once.
"""

from datetime import date, timedelta
import os

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting import ConditionalFormattingList

LIVE = os.environ.get(
    "TASKS_XLSX_LIVE",
    "/Users/kaycschneider/My Drive/STRATEGIC PLANNING/TO DO 4.26.26.xlsx",
)

SAGE_DARK = "6B8E5A"
SAGE_MID = "A8C49A"
SAGE_LIGHT = "E8F0E2"
CREAM = "FAF8F2"
INK = "2E3D2A"
MUTED = "8A8A7E"

ROW_FILL = PatternFill("solid", fgColor=CREAM)
DONE_FILL = PatternFill("solid", fgColor=SAGE_LIGHT)

DAY_PAIRS = [("B","C"), ("D","E"), ("F","G"), ("H","I"), ("J","K"), ("L","M"), ("N","O")]


def current_week_label():
    """Return e.g. 'Apr 27-May 3' for the Mon-Sun week containing today."""
    today = date.today()
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)
    if monday.month == sunday.month:
        return f"{monday.strftime('%b')} {monday.day}-{sunday.day}"
    return f"{monday.strftime('%b')} {monday.day}-{sunday.strftime('%b')} {sunday.day}"


def find_week_tab(wb):
    """Return the live week tab whether it's still 'This Week' or already renamed."""
    if "This Week" in wb.sheetnames:
        return wb["This Week"]
    for name in wb.sheetnames:
        if any(name.startswith(prefix) for prefix in ("Jan ","Feb ","Mar ","Apr ","May ","Jun ","Jul ","Aug ","Sep ","Oct ","Nov ","Dec ")):
            ws = wb[name]
            if ws[f"A2"].value and "WEEK" in str(ws["A2"].value).upper():
                return ws
    return None


def nuke_charts(ws):
    """Remove every chart object from the worksheet."""
    n = len(ws._charts)
    ws._charts = []
    return n


def promote_pct_into_donut_area(ws):
    """Move the per-day % into the merged donut anchor rows 17-21
    and clear the redundant % row 22."""
    big_pct = Font(name="Avenir Next", size=36, bold=True, color=SAGE_DARK)
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for sc, tc in DAY_PAIRS:
        anchor = ws[f"{sc}17"]
        anchor.value = (
            f'=IF(COUNTA({tc}23:{tc}37)=0,"—",'
            f'TEXT(COUNTIF({sc}23:{sc}37,"✅")/COUNTA({tc}23:{tc}37),"0%"))'
        )
        anchor.font = big_pct
        anchor.alignment = center
        anchor.fill = ROW_FILL
        ws[f"{sc}22"].value = ""


def reapply_week_rules(ws):
    """Wipe and re-add conditional formatting on the live week tab."""
    ws.conditional_formatting = ConditionalFormattingList()

    for sc, tc in DAY_PAIRS:
        ws.conditional_formatting.add(
            f"{sc}7:{sc}13",
            FormulaRule(
                formula=[f'={sc}7="✅"'],
                font=Font(name="Avenir Next", size=14, color=SAGE_DARK, bold=True, strike=False),
                fill=DONE_FILL,
            ),
        )
        ws.conditional_formatting.add(
            f"{sc}23:{tc}37",
            FormulaRule(
                formula=[f'=${sc}23="✅"'],
                font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
                fill=DONE_FILL,
            ),
        )


def reapply_todo_rules(ws):
    ws.conditional_formatting = ConditionalFormattingList()
    ws.conditional_formatting.add(
        "B6:G85",
        FormulaRule(
            formula=['=$B6="✅"'],
            font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
            fill=DONE_FILL,
        ),
    )
    ws.conditional_formatting.add(
        "D6:D85",
        FormulaRule(formula=['=$D6="Home"'], fill=PatternFill("solid", fgColor="F4E8D8")),
    )
    ws.conditional_formatting.add(
        "D6:D85",
        FormulaRule(formula=['=$D6="Work"'], fill=PatternFill("solid", fgColor=SAGE_LIGHT)),
    )


def reapply_longterm_rules(ws):
    ws.conditional_formatting = ConditionalFormattingList()
    ws.conditional_formatting.add(
        "B6:F40",
        FormulaRule(
            formula=['=$B6="✅"'],
            font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
            fill=DONE_FILL,
        ),
    )
    ws.conditional_formatting.add(
        "D6:D40",
        FormulaRule(formula=['=$D6="Home"'], fill=PatternFill("solid", fgColor="F4E8D8")),
    )
    ws.conditional_formatting.add(
        "D6:D40",
        FormulaRule(formula=['=$D6="Work"'], fill=PatternFill("solid", fgColor=SAGE_LIGHT)),
    )


def reapply_projects_rules(ws):
    ws.conditional_formatting = ConditionalFormattingList()
    ws.conditional_formatting.add(
        "B6:H25",
        FormulaRule(
            formula=['=$D6="Done"'],
            font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
            fill=DONE_FILL,
        ),
    )
    entity_colors = {
        "Home": "F4E8D8",
        "G&B": "E8F0E2",
        "Myself Renewed": "F5E1E1",
        "Kai Grey": "E1ECF5",
        "Panthera Grey": "ECE1F5",
    }
    for label, color in entity_colors.items():
        ws.conditional_formatting.add(
            "C6:C25",
            FormulaRule(
                formula=[f'=$C6="{label}"'],
                fill=PatternFill("solid", fgColor=color),
            ),
        )


def reapply_gantt_rules(ws, last_milestone_row, last_col_letter, entity_color):
    ws.conditional_formatting = ConditionalFormattingList()
    ws.conditional_formatting.add(
        f"G6:{last_col_letter}{last_milestone_row}",
        FormulaRule(
            formula=['=G6="✅"'],
            fill=PatternFill("solid", fgColor=entity_color),
        ),
    )
    ws.conditional_formatting.add(
        f"B6:F{last_milestone_row}",
        FormulaRule(
            formula=['=$B6="✅"'],
            font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
            fill=DONE_FILL,
        ),
    )


def main():
    wb = load_workbook(LIVE)
    print(f"Loaded {LIVE}")
    print(f"Sheets: {wb.sheetnames}")

    week = find_week_tab(wb)
    if week is None:
        raise SystemExit("Could not find live week tab (looked for 'This Week' or 'Mmm D-D' name)")
    print(f"Live week tab: {week.title!r}")

    n_charts = nuke_charts(week)
    print(f"Removed {n_charts} chart object(s) from {week.title!r}")

    promote_pct_into_donut_area(week)
    print("Promoted per-day % into rows 17-21 anchor; cleared row 22.")

    reapply_week_rules(week)
    print(f"Re-applied conditional formatting on {week.title!r}")

    label = current_week_label()
    if week.title != label:
        old = week.title
        week.title = label
        print(f"Renamed tab: {old!r} -> {label!r}")

    if "To Do" in wb.sheetnames:
        reapply_todo_rules(wb["To Do"])
        print("Re-applied conditional formatting on 'To Do'")

    if "To Do Long Term" in wb.sheetnames:
        reapply_longterm_rules(wb["To Do Long Term"])
        print("Re-applied conditional formatting on 'To Do Long Term'")

    if "Projects" in wb.sheetnames:
        reapply_projects_rules(wb["Projects"])
        print("Re-applied conditional formatting on 'Projects'")

    if "Myself Renewed Healthcare" in wb.sheetnames:
        gantt = wb["Myself Renewed Healthcare"]
        last_milestone_row = 5
        for r in range(6, 30):
            if gantt[f"C{r}"].value:
                last_milestone_row = r
            else:
                break
        last_col_letter = "V"
        for col_idx in range(7, 50):
            from openpyxl.utils import get_column_letter
            letter = get_column_letter(col_idx)
            if gantt[f"{letter}5"].value:
                last_col_letter = letter
            else:
                break
        reapply_gantt_rules(gantt, last_milestone_row, last_col_letter, "F5E1E1")
        print(f"Re-applied conditional formatting on Healthcare Gantt (rows 6-{last_milestone_row}, cols G-{last_col_letter})")

    wb.save(LIVE)
    print(f"\nSaved {LIVE}")


if __name__ == "__main__":
    main()
