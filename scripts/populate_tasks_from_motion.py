"""Populate To Do + To Do Long Term + Projects index in TO DO 4.26.26.xlsx
with Kay's triaged Motion import. Clears the data rows first, then writes
only the approved items. Preserves charts, formulas, conditional formatting,
and styles built by build_tasks_excel.py.
"""

import os
from openpyxl import load_workbook

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(_SCRIPT_DIR)
PATH = os.environ.get(
    "TASKS_XLSX_LIVE",
    os.path.join(_REPO_ROOT, "outputs", "TO DO 4.26.26.xlsx"),
)

# To Do tab — (task, type, project) — project tag links to To Do Long Term or Projects entry
TO_DO = [
    # === HOME (15) ===
    ("Order gift for Lois (Jack Hugh Mackay)", "Home", ""),
    ("Sign up kids for Kano Martial Arts judo class", "Home", ""),
    ("Look into music lessons", "Home", ""),
    ("Find new therapist", "Home", ""),
    ("Sign up for more Italki lessons", "Home", ""),
    ("Fix Kaiden's iPad screen", "Home", ""),
    ("Set up Mohela autopay", "Home", ""),
    ("Fill out buy back form", "Home", ""),
    ("Create travel plans to Paris", "Home", ""),
    ("Download & sign into LeapFrog", "Home", ""),
    ("Renew CitiBike", "Home", ""),
    ("Stool test for microbiome", "Home", ""),
    ("Get assessment vitamin full make up", "Home", ""),
    ("Visit nutritionist", "Home", ""),
    ("Look into GLP1", "Home", ""),

    # === WORK — G&B operations (8) ===
    ("Set up email alerts on FE International", "Work", ""),
    ("Set up email alerts on Benchmark International", "Work", ""),
    ("Switch to higher-yield business savings account", "Work", ""),
    ("Update bank accounts with new G&B address", "Work", ""),
    ("Cancel Linkt by date", "Work", ""),
    ("Cancel Motion & Superhuman annual plans (or downgrade to monthly)", "Work", ""),
    ("File G&B annual report by June", "Work", ""),
    ("Consider signing up for better bookkeeping", "Work", ""),

    # === WORK — Myself Renewed (mapped to Healthcare project) ===
    ("Verify recipient 501(c)(3) status; board vote + grant approval", "Work", "Myself Renewed Healthcare"),
    ("Hold a board meeting + adopt written resolution", "Work", "Myself Renewed Healthcare"),
    ("Add Myself Renewed to QuickBooks (nonprofit category)", "Work", "Myself Renewed Healthcare"),
    ("Cancel Linkt by date", "Work", ""),  # already above; safe to leave dup-checked manually
    ("Confirm registration with NY Charities Bureau", "Work", "Myself Renewed Healthcare"),
    ("Make first personal donation of $1,000", "Work", "Myself Renewed Healthcare"),
    ("Confirm nonprofit bank account exists + PayPal Giving Fund integration", "Work", "Myself Renewed Healthcare"),
    ("Draft personal giving statement aligned with legacy vision", "Work", "Myself Renewed Healthcare"),
    ("Add donation page language + set up donation processor", "Work", "Myself Renewed Healthcare"),

    # === WORK — Other ===
    ("Review Kai Grey Ventures renewal", "Work", ""),
    ("Create brochure for LF", "Work", ""),
]

# Dedupe (Cancel Linkt appeared twice above by mistake — keep first occurrence)
seen = set()
TO_DO_DEDUPED = []
for task, typ, proj in TO_DO:
    key = (task, typ)
    if key in seen:
        continue
    seen.add(key)
    TO_DO_DEDUPED.append((task, typ, proj))
TO_DO = TO_DO_DEDUPED

# To Do Long Term — intents/someday items (item, type, status, notes)
TO_DO_LONG_TERM = [
    ("Create project for French passports", "Home", "Idea", ""),
    ("Plan away with whole Schneider gang", "Home", "Idea", ""),
    ("Marriage Contract — add Aramaic", "Home", "Idea", ""),
    ("Create Wedding Album", "Home", "Idea", ""),
    ("Get engagement ring fixed", "Home", "Idea", ""),
]

# Projects index — (project, entity, status, start, target, tab_link, notes)
PROJECTS_INDEX = [
    (
        "Myself Renewed Healthcare",
        "Myself Renewed",
        "Active",
        "2026-04-27",
        "2026-08-14",
        '=HYPERLINK("#\'Myself Renewed Healthcare\'!A1","Open")',
        "Nonprofit launch + healthcare plan delivery",
    ),
]


def clear_data(ws, first_row, last_row, cols):
    for row in range(first_row, last_row + 1):
        for col in cols:
            ws.cell(row=row, column=col).value = None


def main():
    wb = load_workbook(PATH)

    # ---- To Do ----
    todo = wb["To Do"]
    clear_data(todo, 6, 85, cols=(3, 4, 5, 6, 7))
    for i, (task, typ, proj) in enumerate(TO_DO):
        row = 6 + i
        todo.cell(row=row, column=3, value=task)
        todo.cell(row=row, column=4, value=typ)
        if proj:
            todo.cell(row=row, column=5, value=proj)

    # ---- To Do Long Term ----
    longterm = wb["To Do Long Term"]
    clear_data(longterm, 6, 40, cols=(3, 4, 5, 6))
    for i, (item, typ, status, notes) in enumerate(TO_DO_LONG_TERM):
        row = 6 + i
        longterm.cell(row=row, column=3, value=item)
        longterm.cell(row=row, column=4, value=typ)
        longterm.cell(row=row, column=5, value=status)
        if notes:
            longterm.cell(row=row, column=6, value=notes)

    # ---- Projects index ----
    projects = wb["Projects"]
    clear_data(projects, 6, 25, cols=(2, 3, 4, 5, 6, 7, 8))
    for i, (project, entity, status, start, target, tab_link, notes) in enumerate(PROJECTS_INDEX):
        row = 6 + i
        projects.cell(row=row, column=2, value=project)
        projects.cell(row=row, column=3, value=entity)
        projects.cell(row=row, column=4, value=status)
        projects.cell(row=row, column=5, value=start)
        projects.cell(row=row, column=6, value=target)
        projects.cell(row=row, column=7, value=tab_link)
        projects.cell(row=row, column=8, value=notes)

    wb.save(PATH)
    print(
        f"Wrote {len(TO_DO)} To Do items, "
        f"{len(TO_DO_LONG_TERM)} long-term intents, "
        f"{len(PROJECTS_INDEX)} active project(s) to {PATH}"
    )


if __name__ == "__main__":
    main()
