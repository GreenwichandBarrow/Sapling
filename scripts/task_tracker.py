"""task-tracker-manager skill helper — verbs for Kay's personal task tracker.

Subcommands:
    append                  Add a row to the To Do tab.
    promote                 Move a To Do row into a specific day-slot on the live week tab.
    schedule-to-day-slot    Direct write to a day-slot (no To Do source row required).
    archive                 Sunday rollover ceremony — move live week to a far-right
                            archive tab, rename live to next week, clear data on live.
    archive-todo            Sweep ✅ rows from To Do tab into a running
                            "Completed To Do" tab (created on first run).
    projects-create-gantt   Create a new Gantt project tab cloning the
                            Myself Renewed Healthcare structure; updates Projects index.
    reformat                Re-apply conditional formatting + fix donuts.
    report                  Markdown health summary (overdue, empty slots, carryover).
    gantt-tick              Fill a week-cell on a Gantt project tab.

Always backs up the live file before writing. Refuses to write while Excel
has the file open (would clobber Excel's autosave or vice versa).
"""

from __future__ import annotations

import argparse
import os
import shutil
import subprocess
import sys
from copy import copy
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(_SCRIPT_DIR)
LIVE = os.environ.get(
    "TASKS_XLSX_LIVE",
    os.path.join(_REPO_ROOT, "outputs", "TO DO 4.26.26.xlsx"),
)
BACKUP_DIR = Path(LIVE).parent
BACKUP_PREFIX = Path(LIVE).stem + ".bak."
BACKUP_KEEP = 5

SAGE_DARK = "6B8E5A"
SAGE_MID = "A8C49A"
SAGE_LIGHT = "E8F0E2"
CREAM = "FAF8F2"
INK = "2E3D2A"
MUTED = "8A8A7E"

ROW_FILL = PatternFill("solid", fgColor=CREAM)
DONE_FILL = PatternFill("solid", fgColor=SAGE_LIGHT)

DAY_PAIRS = [("B","C"), ("D","E"), ("F","G"), ("H","I"), ("J","K"), ("L","M"), ("N","O")]
DAY_BY_NAME = {
    "mon": 0, "monday": 0,
    "tue": 1, "tuesday": 1,
    "wed": 2, "wednesday": 2,
    "thu": 3, "thursday": 3,
    "fri": 4, "friday": 4,
    "sat": 5, "saturday": 5,
    "sun": 6, "sunday": 6,
}


# ---------------------------------------------------------------- shared utils

def excel_has_open(path: str) -> bool:
    """Return True if Excel currently has a handle on the file."""
    try:
        out = subprocess.run(["lsof", path], capture_output=True, text=True, timeout=5)
        return "Microsoft" in out.stdout or "Excel" in out.stdout
    except Exception:
        return False


def assert_writable(path: str) -> None:
    if excel_has_open(path):
        sys.exit(f"task-tracker-manager: refused — Excel has {path!r} open. Close it first (Cmd+Q).")
    if not Path(path).exists():
        sys.exit(f"task-tracker-manager: refused — file not found at {path!r}")


def backup(path: str) -> str:
    from datetime import datetime
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    dst = BACKUP_DIR / f"{BACKUP_PREFIX}{ts}.xlsx"
    shutil.copy2(path, dst)
    backups = sorted(BACKUP_DIR.glob(f"{BACKUP_PREFIX}*.xlsx"))
    for old in backups[:-BACKUP_KEEP]:
        old.unlink()
    return str(dst)


def current_week_label(today: date | None = None) -> str:
    if today is None:
        today = date.today()
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)
    if monday.month == sunday.month:
        return f"{monday.strftime('%b')} {monday.day}-{sunday.day}"
    return f"{monday.strftime('%b')} {monday.day}-{sunday.strftime('%b')} {sunday.day}"


def find_week_tab(wb):
    if "This Week" in wb.sheetnames:
        return wb["This Week"]
    months = ("Jan ","Feb ","Mar ","Apr ","May ","Jun ","Jul ","Aug ","Sep ","Oct ","Nov ","Dec ")
    for name in wb.sheetnames:
        if any(name.startswith(m) for m in months) and not name.startswith("archive_"):
            return wb[name]
    return None


def trace(verb: str, slug: str, lines: list[str]) -> None:
    today = date.today().isoformat()
    repo = Path(__file__).resolve().parent.parent
    trace_dir = repo / "brain" / "traces"
    trace_dir.mkdir(parents=True, exist_ok=True)
    trace_path = trace_dir / f"{today}-task-tracker-{verb}-{slug}.md"
    body = "\n".join([
        "---",
        f"name: task-tracker {verb} — {slug}",
        f"date: {today}",
        f"type: trace",
        f"tags:",
        f"  - date/{today}",
        f"  - trace",
        f"  - skill/task-tracker-manager",
        f"  - verb/{verb}",
        "---",
        f"# task-tracker {verb} — {slug}",
        "",
        *lines,
    ])
    trace_path.write_text(body)


# ------------------------------------------------------------------- verbs

def cmd_append(args) -> int:
    if args.type not in ("Work", "Home"):
        sys.exit("task-tracker-manager: --type must be Work or Home")

    assert_writable(LIVE)
    bk = backup(LIVE)
    wb = load_workbook(LIVE)
    if "To Do" not in wb.sheetnames:
        sys.exit("task-tracker-manager: 'To Do' tab missing")

    ws = wb["To Do"]
    target_row = None
    for r in range(6, 86):
        if not ws[f"C{r}"].value:
            target_row = r
            break
    if target_row is None:
        sys.exit("task-tracker-manager: To Do tab is full (rows 6-85). Time to widen capacity.")

    ws[f"B{target_row}"] = "☐"
    ws[f"C{target_row}"] = args.task
    ws[f"D{target_row}"] = args.type
    if args.project:
        ws[f"E{target_row}"] = args.project
    if args.due:
        ws[f"F{target_row}"] = args.due
    if args.notes:
        ws[f"G{target_row}"] = args.notes

    wb.save(LIVE)
    trace("append", args.task[:40].lower().replace(" ", "-"), [
        f"- task: {args.task}",
        f"- type: {args.type}",
        f"- project: {args.project or '—'}",
        f"- due: {args.due or '—'}",
        f"- row: {target_row}",
        f"- backup: {bk}",
        f"- rollback: cp {bk!r} {LIVE!r}",
    ])
    print(f'task-tracker-manager: appended row {target_row} ("{args.task}" / {args.type} / {args.project or "—"} / {args.due or "—"})')
    return 0


def cmd_promote(args) -> int:
    day_idx = DAY_BY_NAME.get(args.day.lower())
    if day_idx is None:
        sys.exit(f"task-tracker-manager: unknown day {args.day!r}. Use Mon..Sun.")
    if not (1 <= args.slot <= 15):
        sys.exit("task-tracker-manager: --slot must be 1..15")

    assert_writable(LIVE)
    bk = backup(LIVE)
    wb = load_workbook(LIVE)

    if "To Do" not in wb.sheetnames:
        sys.exit("task-tracker-manager: 'To Do' tab missing")
    todo = wb["To Do"]
    task_text = todo[f"C{args.todo_row}"].value
    if not task_text:
        sys.exit(f"task-tracker-manager: To Do row {args.todo_row} is empty")

    week = find_week_tab(wb)
    if week is None:
        sys.exit("task-tracker-manager: live week tab not found")

    sc, tc = DAY_PAIRS[day_idx]
    priority_row = 22 + args.slot
    existing = week[f"{tc}{priority_row}"].value
    if existing:
        sys.exit(f'task-tracker-manager: refused promote — {args.day} slot {args.slot} already contains "{existing}"')

    week[f"{tc}{priority_row}"] = task_text
    todo[f"B{args.todo_row}"] = "→"

    wb.save(LIVE)
    trace("promote", f"{args.day.lower()}-{args.slot}", [
        f"- todo_row: {args.todo_row}",
        f"- task: {task_text}",
        f"- promoted_to: {week.title} {sc}/{tc} row {priority_row} ({args.day} slot {args.slot})",
        f"- backup: {bk}",
        f"- rollback: cp {bk!r} {LIVE!r}",
    ])
    print(f'task-tracker-manager: promoted To Do row {args.todo_row} → {week.title} {args.day} slot {args.slot} ("{task_text}")')
    return 0


def cmd_archive(args) -> int:
    """Sunday rollover: clone live week tab to a visible archive tab parked far-right,
    rename live to the upcoming week, clear data on live."""
    assert_writable(LIVE)
    bk = backup(LIVE)
    wb = load_workbook(LIVE)

    week = find_week_tab(wb)
    if week is None:
        sys.exit("task-tracker-manager: live week tab not found")
    old_label = week.title

    archive_name = f"archive_{old_label}"
    suffix = 1
    while archive_name in wb.sheetnames:
        suffix += 1
        archive_name = f"archive_{old_label}_v{suffix}"

    src = week
    dst = wb.copy_worksheet(src)
    dst.title = archive_name
    dst.sheet_state = "visible"
    # Park the archive tab at the far right so the live tab + reference tabs stay easy to find.
    current_idx = wb.sheetnames.index(dst.title)
    wb.move_sheet(dst.title, offset=len(wb.sheetnames) - 1 - current_idx)

    today = date.today()
    # Monday edge case: when run on Monday, the new live tab is THIS week, not next.
    # Sunday weekday=6 → +1 = next Monday. Mon weekday=0 → +0 = today. Tue..Sat → next Mon.
    if today.weekday() == 0:
        new_monday = today
    else:
        new_monday = today + timedelta(days=(7 - today.weekday()))
    new_label = current_week_label(new_monday)
    src.title = new_label

    for r in range(7, 14):
        for sc, tc in DAY_PAIRS:
            src[f"{sc}{r}"] = "☐"
    for r in range(23, 38):
        for sc, tc in DAY_PAIRS:
            src[f"{sc}{r}"] = "☐"
            src[f"{tc}{r}"] = ""
    for r in range(40, 48):
        for sc, tc in DAY_PAIRS:
            src[f"{sc}{r}"] = ""

    wb.save(LIVE)
    trace("archive", old_label.lower().replace(" ", "-"), [
        f"- archived: {old_label} → visible tab {archive_name} (parked far-right)",
        f"- live tab renamed: {old_label} → {new_label}",
        f"- cleared: habits (rows 7-13), priorities (rows 23-37), notes (rows 40-47)",
        f"- backup: {bk}",
        f"- rollback: cp {bk!r} {LIVE!r}",
    ])
    print(f'task-tracker-manager: archived "{old_label}" → visible far-right tab {archive_name}; live tab now "{new_label}"')
    return 0


# ---------------------------------------------------------------- archive-todo

COMPLETED_TODO_TAB = "Completed To Do"


def _ensure_completed_todo_tab(wb):
    """Create the 'Completed To Do' tab if missing. Returns the worksheet."""
    if COMPLETED_TODO_TAB in wb.sheetnames:
        return wb[COMPLETED_TODO_TAB]

    ws = wb.create_sheet(title=COMPLETED_TODO_TAB)

    # Column widths — mirror To Do (B narrow status, C wide task, D-H standard)
    widths = {"A": 2, "B": 4, "C": 60, "D": 10, "E": 18, "F": 12, "G": 32, "H": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    title_font = Font(name="Avenir Next", size=18, bold=True, color=SAGE_DARK)
    sub_font = Font(name="Avenir Next", size=10, italic=True, color=MUTED)
    header_font = Font(name="Avenir Next", size=10, bold=True, color=INK)
    header_fill = PatternFill("solid", fgColor=SAGE_LIGHT)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws["B2"] = "COMPLETED TO DO  ·  running list of done items"
    ws["B2"].font = title_font
    ws["B2"].alignment = left
    ws.merge_cells("B2:H2")

    ws["B3"] = "Auto-populated when ✅ rows sweep out of the To Do tab. Most recent at top."
    ws["B3"].font = sub_font
    ws["B3"].alignment = left
    ws.merge_cells("B3:H3")

    headers = [("B5", ""), ("C5", "Task"), ("D5", "Type"), ("E5", "Project"),
               ("F5", "Due"), ("G5", "Notes"), ("H5", "Completed")]
    for ref, val in headers:
        ws[ref] = val
        ws[ref].font = header_font
        ws[ref].fill = header_fill
        ws[ref].alignment = center if ref in ("B5", "H5") else left

    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[5].height = 20

    # Conditional formatting: dim completed rows (they all are, but consistent with To Do)
    ws.conditional_formatting = ConditionalFormattingList()
    ws.conditional_formatting.add("B6:H200", FormulaRule(
        formula=['=$B6="✅"'],
        font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
        fill=DONE_FILL,
    ))
    ws.conditional_formatting.add("D6:D200", FormulaRule(
        formula=['=$D6="Home"'], fill=PatternFill("solid", fgColor="F4E8D8")))
    ws.conditional_formatting.add("D6:D200", FormulaRule(
        formula=['=$D6="Work"'], fill=PatternFill("solid", fgColor=SAGE_LIGHT)))

    return ws


def cmd_archive_todo(args) -> int:
    """Sweep ✅ rows from To Do tab → Completed To Do tab (created on first run).
    Most-recent at top: new completions land at row 6, existing rows shift down."""
    assert_writable(LIVE)
    bk = backup(LIVE)
    wb = load_workbook(LIVE)

    if "To Do" not in wb.sheetnames:
        sys.exit("task-tracker-manager: 'To Do' tab missing")
    src = wb["To Do"]

    dst = _ensure_completed_todo_tab(wb)
    completed_today = date.today().isoformat()

    # Collect ✅ rows from To Do
    swept = []
    for r in range(6, src.max_row + 1):
        if src[f"B{r}"].value == "✅" and src[f"C{r}"].value:
            swept.append({
                "row": r,
                "task": src[f"C{r}"].value,
                "type": src[f"D{r}"].value or "",
                "project": src[f"E{r}"].value or "",
                "due": src[f"F{r}"].value or "",
                "notes": src[f"G{r}"].value or "",
            })

    if not swept:
        # Still save: the _ensure_completed_todo_tab call above may have just created the tab.
        wb.save(LIVE)
        msg = "tab created" if dst.max_row <= 5 else "tab present"
        print(f"task-tracker-manager: archive-todo — no ✅ rows in To Do to sweep ({msg})")
        return 0

    # Find first empty row in Completed To Do (data starts at row 6)
    insert_at = 6
    for r in range(6, dst.max_row + 2):
        if not dst[f"C{r}"].value:
            insert_at = r
            break

    # Append swept rows
    for i, item in enumerate(swept):
        r = insert_at + i
        dst[f"B{r}"] = "✅"
        dst[f"C{r}"] = item["task"]
        dst[f"D{r}"] = item["type"]
        dst[f"E{r}"] = item["project"]
        dst[f"F{r}"] = str(item["due"])[:10] if item["due"] else ""
        dst[f"G{r}"] = item["notes"]
        dst[f"H{r}"] = completed_today

    # Clear swept rows from To Do (status, task, type, project, due, notes)
    for item in swept:
        r = item["row"]
        for col in ("B", "C", "D", "E", "F", "G"):
            src[f"{col}{r}"] = ""

    wb.save(LIVE)
    slug = f"sweep-{len(swept)}"
    trace("archive-todo", slug, [
        f"- swept: {len(swept)} ✅ rows from To Do",
        f"- destination: '{COMPLETED_TODO_TAB}' rows {insert_at}..{insert_at + len(swept) - 1}",
        f"- completed_date: {completed_today}",
        f"- backup: {bk}",
        f"- rollback: cp {bk!r} {LIVE!r}",
    ])
    print(f"task-tracker-manager: archive-todo swept {len(swept)} ✅ row(s) → '{COMPLETED_TODO_TAB}' rows {insert_at}..{insert_at + len(swept) - 1}")
    return 0


# ---------------------------------------------------------- schedule-to-day-slot

def cmd_schedule_to_day_slot(args) -> int:
    """Write a task directly into a day-slot on the live week tab (no To Do source)."""
    day_idx = DAY_BY_NAME.get(args.day.lower())
    if day_idx is None:
        sys.exit(f"task-tracker-manager: unknown day {args.day!r}. Use Mon..Sun.")

    assert_writable(LIVE)
    bk = backup(LIVE)
    wb = load_workbook(LIVE)

    week = find_week_tab(wb)
    if week is None:
        sys.exit("task-tracker-manager: live week tab not found")

    sc, tc = DAY_PAIRS[day_idx]

    # Resolve slot — explicit, or auto-pick first empty
    if args.slot is not None:
        if not (1 <= args.slot <= 15):
            sys.exit("task-tracker-manager: --slot must be 1..15")
        slot = args.slot
    else:
        slot = None
        for s in range(1, 16):
            if not week[f"{tc}{22 + s}"].value:
                slot = s
                break
        if slot is None:
            sys.exit(f"task-tracker-manager: refused schedule-to-day-slot — {args.day} has no empty slots")

    priority_row = 22 + slot
    existing = week[f"{tc}{priority_row}"].value
    if existing and not args.force:
        sys.exit(
            f'task-tracker-manager: refused schedule-to-day-slot — {args.day} slot {slot} '
            f'already contains "{existing}" (use --force to overwrite)'
        )

    week[f"{tc}{priority_row}"] = args.task
    week[f"{sc}{priority_row}"] = "☐"

    wb.save(LIVE)
    trace("schedule-to-day-slot", f"{args.day.lower()}-{slot}", [
        f"- task: {args.task}",
        f"- placement: {week.title} {sc}/{tc} row {priority_row} ({args.day} slot {slot})",
        f"- overwrote: {existing!r}" if existing else "- overwrote: (slot was empty)",
        f"- backup: {bk}",
        f"- rollback: cp {bk!r} {LIVE!r}",
    ])
    print(f'task-tracker-manager: scheduled "{args.task}" → {week.title} {args.day} slot {slot}')
    return 0


# ---------------------------------------------------------- projects-create-gantt

ENTITY_COLORS = {
    "Home": "F4E8D8",
    "G&B": "E8F0E2",
    "Myself Renewed": "F5E1E1",
    "Kai Grey": "E1ECF5",
    "Panthera Grey": "ECE1F5",
}


def _weekly_labels(start_iso: str, n_weeks: int) -> list[str]:
    """Return n_weeks weekly Monday labels in M/D format starting from the Monday of start_iso."""
    y, m, d = (int(x) for x in start_iso.split("-"))
    start = date(y, m, d)
    monday = start - timedelta(days=start.weekday())
    out = []
    for i in range(n_weeks):
        wk = monday + timedelta(days=7 * i)
        out.append(f"{wk.month}/{wk.day}")
    return out


def cmd_projects_create_gantt(args) -> int:
    """Create a new Gantt project tab (clones Myself Renewed Healthcare structure).
    Adds/updates a row in the Projects index with a HYPERLINK to the new tab."""
    if len(args.project) > 31 or any(ch in args.project for ch in ":\\/?*[]"):
        sys.exit(f"task-tracker-manager: invalid tab name {args.project!r} (Excel rules: ≤31 chars, no :\\/?*[])")

    assert_writable(LIVE)
    bk = backup(LIVE)
    wb = load_workbook(LIVE)

    if args.project in wb.sheetnames:
        sys.exit(f"task-tracker-manager: tab {args.project!r} already exists — pick a different name or delete the existing tab first")

    if "Projects" not in wb.sheetnames:
        sys.exit("task-tracker-manager: 'Projects' index tab missing — cannot wire up new Gantt tab")

    n_weeks = args.weeks
    if n_weeks < 4 or n_weeks > 30:
        sys.exit("task-tracker-manager: --weeks must be 4..30")

    week_labels = _weekly_labels(args.start, n_weeks)
    entity_color = ENTITY_COLORS.get(args.entity, "E8F0E2")  # default to G&B sage

    # ----- 1) Create the Gantt tab
    gantt = wb.create_sheet(title=args.project)

    title_font = Font(name="Avenir Next", size=18, bold=True, color=SAGE_DARK)
    sub_font = Font(name="Avenir Next", size=10, italic=True, color=MUTED)
    header_font = Font(name="Avenir Next", size=10, bold=True, color=INK)
    header_fill = PatternFill("solid", fgColor=SAGE_LIGHT)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    last_col_idx = 6 + n_weeks  # F=6, then n_weeks of week columns
    last_col_letter = get_column_letter(last_col_idx)

    gantt["B2"] = args.project.upper()
    gantt["B2"].font = title_font
    gantt["B2"].alignment = left
    gantt.merge_cells(f"B2:{last_col_letter}2")

    gantt["B3"] = (
        f"Entity: {args.entity}  ·  Tick the week boxes you're actively working on "
        f"each milestone — the row builds into a Gantt bar"
    )
    gantt["B3"].font = sub_font
    gantt["B3"].alignment = left
    gantt.merge_cells(f"B3:{last_col_letter}3")

    # Header row 5: C=Milestone, D=Start, E=Target, F=Notes, G..=week labels
    headers = [("C5", "Milestone"), ("D5", "Start"), ("E5", "Target"), ("F5", "Notes")]
    for ref, val in headers:
        gantt[ref] = val
        gantt[ref].font = header_font
        gantt[ref].fill = header_fill
        gantt[ref].alignment = left
    for i, label in enumerate(week_labels):
        col = get_column_letter(7 + i)
        gantt[f"{col}5"] = label
        gantt[f"{col}5"].font = header_font
        gantt[f"{col}5"].fill = header_fill
        gantt[f"{col}5"].alignment = center

    # Blank milestone scaffold: 10 empty rows (6..15), status checkbox in B, week boxes ☐
    for r in range(6, 16):
        gantt[f"B{r}"] = "☐"
        gantt[f"B{r}"].alignment = center
        for i in range(n_weeks):
            col = get_column_letter(7 + i)
            gantt[f"{col}{r}"] = "☐"
            gantt[f"{col}{r}"].alignment = center

    # Column widths
    gantt.column_dimensions["A"].width = 2
    gantt.column_dimensions["B"].width = 4
    gantt.column_dimensions["C"].width = 50
    gantt.column_dimensions["D"].width = 12
    gantt.column_dimensions["E"].width = 12
    gantt.column_dimensions["F"].width = 30
    for i in range(n_weeks):
        gantt.column_dimensions[get_column_letter(7 + i)].width = 6
    gantt.row_dimensions[2].height = 28
    gantt.row_dimensions[3].height = 18
    gantt.row_dimensions[5].height = 20

    # Conditional formatting — week box ✅ → entity-color fill (the Gantt bar)
    gantt.conditional_formatting = ConditionalFormattingList()
    gantt.conditional_formatting.add(f"G6:{last_col_letter}15", FormulaRule(
        formula=['=G6="✅"'],
        fill=PatternFill("solid", fgColor=entity_color),
    ))
    gantt.conditional_formatting.add(f"B6:F15", FormulaRule(
        formula=['=$B6="✅"'],
        font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
        fill=DONE_FILL,
    ))

    # ----- 2) Update Projects index
    pj = wb["Projects"]
    existing_row = None
    for r in range(6, pj.max_row + 2):
        if pj[f"B{r}"].value == args.project:
            existing_row = r
            break

    hyperlink_formula = f'=HYPERLINK("#\'{args.project}\'!A1","Open")'

    if existing_row is None:
        # Append a new row
        target_row = 6
        for r in range(6, pj.max_row + 2):
            if not pj[f"B{r}"].value:
                target_row = r
                break
        pj[f"B{target_row}"] = args.project
        pj[f"C{target_row}"] = args.entity
        pj[f"D{target_row}"] = args.status
        pj[f"E{target_row}"] = args.start
        pj[f"F{target_row}"] = args.target
        pj[f"G{target_row}"] = hyperlink_formula
        if args.notes:
            pj[f"H{target_row}"] = args.notes
        project_row_msg = f"appended at row {target_row}"
    else:
        # Update the existing row's Tab hyperlink (and status if requested)
        pj[f"G{existing_row}"] = hyperlink_formula
        if args.status:
            pj[f"D{existing_row}"] = args.status
        project_row_msg = f"updated row {existing_row} (existing entry)"

    wb.save(LIVE)
    slug = args.project.lower().replace(" ", "-")
    trace("projects-create-gantt", slug, [
        f"- project: {args.project}",
        f"- entity: {args.entity} (color #{entity_color})",
        f"- gantt tab: created with {n_weeks} weekly columns from {args.start}",
        f"- projects index: {project_row_msg}",
        f"- backup: {bk}",
        f"- rollback: cp {bk!r} {LIVE!r}",
    ])
    print(f'task-tracker-manager: created Gantt tab "{args.project}" ({n_weeks} weeks from {args.start}); Projects index {project_row_msg}')
    return 0


def cmd_reformat(args) -> int:
    """Re-apply conditional formatting + fix donut layout. Idempotent."""
    assert_writable(LIVE)
    bk = backup(LIVE)
    wb = load_workbook(LIVE)
    week = find_week_tab(wb)
    if week is None:
        sys.exit("task-tracker-manager: live week tab not found")

    n_charts = len(week._charts)
    week._charts = []

    big_pct = Font(name="Avenir Next", size=36, bold=True, color=SAGE_DARK)
    center = Alignment(horizontal="center", vertical="center")
    for sc, tc in DAY_PAIRS:
        anchor = week[f"{sc}17"]
        anchor.value = (
            f'=IF(COUNTA({tc}23:{tc}37)=0,"—",'
            f'TEXT(COUNTIF({sc}23:{sc}37,"✅")/COUNTA({tc}23:{tc}37),"0%"))'
        )
        anchor.font = big_pct
        anchor.alignment = center
        anchor.fill = ROW_FILL
        week[f"{sc}22"].value = ""

    week.conditional_formatting = ConditionalFormattingList()
    for sc, tc in DAY_PAIRS:
        week.conditional_formatting.add(f"{sc}7:{sc}13", FormulaRule(
            formula=[f'={sc}7="✅"'],
            font=Font(name="Avenir Next", size=14, color=SAGE_DARK, bold=True),
            fill=DONE_FILL,
        ))
        week.conditional_formatting.add(f"{sc}23:{tc}37", FormulaRule(
            formula=[f'=${sc}23="✅"'],
            font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
            fill=DONE_FILL,
        ))

    if "To Do" in wb.sheetnames:
        td = wb["To Do"]
        td.conditional_formatting = ConditionalFormattingList()
        td.conditional_formatting.add("B6:G85", FormulaRule(
            formula=['=$B6="✅"'],
            font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
            fill=DONE_FILL,
        ))
        td.conditional_formatting.add("D6:D85", FormulaRule(
            formula=['=$D6="Home"'], fill=PatternFill("solid", fgColor="F4E8D8")))
        td.conditional_formatting.add("D6:D85", FormulaRule(
            formula=['=$D6="Work"'], fill=PatternFill("solid", fgColor=SAGE_LIGHT)))

    if "To Do Long Term" in wb.sheetnames:
        lt = wb["To Do Long Term"]
        lt.conditional_formatting = ConditionalFormattingList()
        lt.conditional_formatting.add("B6:F40", FormulaRule(
            formula=['=$B6="✅"'],
            font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
            fill=DONE_FILL,
        ))
        lt.conditional_formatting.add("D6:D40", FormulaRule(
            formula=['=$D6="Home"'], fill=PatternFill("solid", fgColor="F4E8D8")))
        lt.conditional_formatting.add("D6:D40", FormulaRule(
            formula=['=$D6="Work"'], fill=PatternFill("solid", fgColor=SAGE_LIGHT)))

    if "Projects" in wb.sheetnames:
        pj = wb["Projects"]
        pj.conditional_formatting = ConditionalFormattingList()
        pj.conditional_formatting.add("B6:H25", FormulaRule(
            formula=['=$D6="Done"'],
            font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
            fill=DONE_FILL,
        ))
        for label, color in [
            ("Home", "F4E8D8"), ("G&B", "E8F0E2"), ("Myself Renewed", "F5E1E1"),
            ("Kai Grey", "E1ECF5"), ("Panthera Grey", "ECE1F5"),
        ]:
            pj.conditional_formatting.add("C6:C25", FormulaRule(
                formula=[f'=$C6="{label}"'],
                fill=PatternFill("solid", fgColor=color),
            ))

    if "Myself Renewed Healthcare" in wb.sheetnames:
        gantt = wb["Myself Renewed Healthcare"]
        last_row = 5
        for r in range(6, 30):
            if gantt[f"C{r}"].value:
                last_row = r
            else:
                break
        last_col_letter = "V"
        for col_idx in range(7, 50):
            letter = get_column_letter(col_idx)
            if gantt[f"{letter}5"].value:
                last_col_letter = letter
            else:
                break
        gantt.conditional_formatting = ConditionalFormattingList()
        gantt.conditional_formatting.add(f"G6:{last_col_letter}{last_row}", FormulaRule(
            formula=['=G6="✅"'],
            fill=PatternFill("solid", fgColor="F5E1E1"),
        ))
        gantt.conditional_formatting.add(f"B6:F{last_row}", FormulaRule(
            formula=['=$B6="✅"'],
            font=Font(name="Avenir Next", size=10, color=MUTED, strike=True),
            fill=DONE_FILL,
        ))

    label = current_week_label()
    renamed = False
    if week.title == "This Week":
        week.title = label
        renamed = True

    wb.save(LIVE)
    print(f"task-tracker-manager: reformatted ({n_charts} chart(s) removed; tab {'renamed to ' + label if renamed else 'kept as ' + week.title})")
    return 0


def cmd_report(args) -> int:
    wb = load_workbook(LIVE, data_only=True)
    today = date.today()
    today_iso = today.isoformat()

    overdue = []
    unscheduled = []
    if "To Do" in wb.sheetnames:
        td = wb["To Do"]
        for r in range(6, 86):
            status = td[f"B{r}"].value
            task = td[f"C{r}"].value
            due = td[f"F{r}"].value
            if not task:
                continue
            if status == "✅":
                continue
            if due:
                due_str = str(due)[:10]
                if due_str < today_iso:
                    overdue.append(f"  - row {r}: {task} (due {due_str})")
            else:
                unscheduled.append(f"  - row {r}: {task}")

    week = find_week_tab(wb)
    empty_slots = []
    if week is not None:
        tomorrow_idx = (today.weekday() + 1) % 7
        sc, tc = DAY_PAIRS[tomorrow_idx]
        empty_count = 0
        for r in range(23, 38):
            if not week[f"{tc}{r}"].value:
                empty_count += 1
        empty_slots.append(f"  - tomorrow ({['Mon','Tue','Wed','Thu','Fri','Sat','Sun'][tomorrow_idx]}): {empty_count}/15 empty")

    lines = [f"## Tracker health ({today_iso})", ""]
    lines.append(f"**Overdue ({len(overdue)}):**")
    lines.extend(overdue if overdue else ["  - none"])
    lines.append("")
    lines.append(f"**Unscheduled / no due date ({len(unscheduled)}):**")
    lines.extend(unscheduled[:10] if unscheduled else ["  - none"])
    if len(unscheduled) > 10:
        lines.append(f"  - ... and {len(unscheduled)-10} more")
    lines.append("")
    lines.append("**Priority slot capacity:**")
    lines.extend(empty_slots)

    print("\n".join(lines))
    return 0


def cmd_gantt_tick(args) -> int:
    assert_writable(LIVE)
    bk = backup(LIVE)
    wb = load_workbook(LIVE)
    if args.project not in wb.sheetnames:
        sys.exit(f"task-tracker-manager: project tab {args.project!r} not found")
    ws = wb[args.project]
    cell_ref = f"{args.week_col}{args.milestone_row}"
    ws[cell_ref] = "✅"
    wb.save(LIVE)
    trace("gantt-tick", f"{args.project.lower().replace(' ', '-')}-{cell_ref.lower()}", [
        f"- project: {args.project}",
        f"- cell: {cell_ref}",
        f"- backup: {bk}",
        f"- rollback: cp {bk!r} {LIVE!r}",
    ])
    print(f'task-tracker-manager: ticked {args.project} {cell_ref}')
    return 0


# --------------------------------------------------------------------- main

def main():
    p = argparse.ArgumentParser(prog="task_tracker")
    sub = p.add_subparsers(dest="cmd", required=True)

    a = sub.add_parser("append")
    a.add_argument("--task", required=True)
    a.add_argument("--type", required=True, choices=["Work", "Home"])
    a.add_argument("--project", default="")
    a.add_argument("--due", default="")
    a.add_argument("--notes", default="")
    a.set_defaults(func=cmd_append)

    pr = sub.add_parser("promote")
    pr.add_argument("--todo-row", type=int, required=True)
    pr.add_argument("--day", required=True)
    pr.add_argument("--slot", type=int, required=True)
    pr.set_defaults(func=cmd_promote)

    ar = sub.add_parser("archive")
    ar.set_defaults(func=cmd_archive)

    at = sub.add_parser("archive-todo")
    at.set_defaults(func=cmd_archive_todo)

    sds = sub.add_parser("schedule-to-day-slot")
    sds.add_argument("--task", required=True)
    sds.add_argument("--day", required=True)
    sds.add_argument("--slot", type=int, default=None,
                     help="1..15; if omitted, auto-pick first empty slot for the day")
    sds.add_argument("--force", action="store_true",
                     help="overwrite the slot even if occupied")
    sds.set_defaults(func=cmd_schedule_to_day_slot)

    pcg = sub.add_parser("projects-create-gantt")
    pcg.add_argument("--project", required=True, help="Tab name (≤31 chars, no :\\/?*[])")
    pcg.add_argument("--entity", required=True,
                     help="Home | G&B | Myself Renewed | Kai Grey | Panthera Grey")
    pcg.add_argument("--status", default="Active",
                     help="Status column on Projects index (default: Active)")
    pcg.add_argument("--start", required=True, help="ISO date YYYY-MM-DD")
    pcg.add_argument("--target", required=True, help="ISO date YYYY-MM-DD")
    pcg.add_argument("--weeks", type=int, default=16,
                     help="number of weekly columns (4..30; default 16)")
    pcg.add_argument("--notes", default="")
    pcg.set_defaults(func=cmd_projects_create_gantt)

    rf = sub.add_parser("reformat")
    rf.set_defaults(func=cmd_reformat)

    rp = sub.add_parser("report")
    rp.set_defaults(func=cmd_report)

    gt = sub.add_parser("gantt-tick")
    gt.add_argument("--project", required=True)
    gt.add_argument("--milestone-row", type=int, required=True)
    gt.add_argument("--week-col", required=True)
    gt.set_defaults(func=cmd_gantt_tick)

    args = p.parse_args()
    sys.exit(args.func(args))


if __name__ == "__main__":
    main()
